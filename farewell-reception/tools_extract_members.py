# 送別会Excelから参加者データを抽出し、D1インポート用のJSON/SQLを作る
# ※個人情報のためリポジトリにはコミットしない（scratch内で生成）
import sys, json
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

wb = load_workbook("miyamoto.xlsx", data_only=True)
ws = wb["収支管理"]
wss = wb["設定"]

event = {
    "name": wss["C5"].value,
    "date_serial": wss["C6"].value,
    "venue": wss["C8"].value,
    "address": wss["C9"].value,
    "url": wss["C10"].value,
    "organizer": wss["C7"].value,
}

members = []
for r in range(7, 47):
    nm = ws[f"C{r}"].value
    if not nm:
        continue
    fee = ws[f"E{r}"].value or 0
    sup = ws[f"F{r}"].value or 0
    members.append({
        "row": r,
        "name": str(nm).strip(),
        "rank": ws[f"D{r}"].value,
        "fee": int(fee),
        "support": int(sup),
        "due": int(fee) + int(sup),        # 当日徴収する金額（会費＋ご支援金）
        "alcohol": ws[f"H{r}"].value,
        "meal_limit": ws[f"I{r}"].value,
        "shuttle": ws[f"J{r}"].value,
        "note": ws[f"K{r}"].value or "",
    })

print(f"参加者 {len(members)}名")
total = sum(m["due"] for m in members)
print(f"徴収予定合計: {total:,}円")
print(f"招待(徴収0): {sum(1 for m in members if m['due']==0)}名")

# 部署をname先頭から分離（例: '生技 山田' -> 部署'生技' 氏名'山田'）
for m in members:
    parts = m["name"].split(None, 1)
    if len(parts) == 2 and len(parts[0]) <= 3:
        m["dept"], m["person"] = parts[0], parts[1]
    else:
        m["dept"], m["person"] = "", m["name"]

with open("members.json", "w", encoding="utf-8") as f:
    json.dump({"event": event, "members": members}, f, ensure_ascii=False, indent=2)
print("→ members.json 出力")

# D1 INSERT文
def esc(s):
    return str(s).replace("'", "''") if s is not None else ""

lines = ["-- 参加者データ（個人情報：リポジトリにコミットしないこと）"]
for i, m in enumerate(members, 1):
    lines.append(
        "INSERT INTO attendees (id, dept, name, rank, fee, support, due, alcohol, shuttle, note) VALUES "
        f"({i}, '{esc(m['dept'])}', '{esc(m['person'])}', '{esc(m['rank'])}', {m['fee']}, {m['support']}, {m['due']}, "
        f"'{esc(m['alcohol'])}', '{esc(m['shuttle'])}', '{esc(m['note'])}');"
    )
with open("seed_attendees.sql", "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")
print("→ seed_attendees.sql 出力")

print("\n--- 徴収額の分布 ---")
from collections import Counter
for due, cnt in sorted(Counter(m["due"] for m in members).items()):
    print(f"  {due:,}円 × {cnt}名")

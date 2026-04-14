# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

wb = Workbook()
ws = wb.active
ws.title = '会費収支報告書'

F  = lambda **kw: Font(name='游ゴシック', **kw)
f_title = F(size=14, bold=True)
f_hdr   = F(size=10, bold=True, color='FFFFFF')
f_body  = F(size=10)
f_bold  = F(size=10, bold=True)
f_sum   = F(size=11, bold=True)
f_big   = F(size=12, bold=True)
f_note  = F(size=9,  italic=True, color='666666')
f_sub   = F(size=9,  color='555555', italic=True)
f_red   = F(size=10, bold=True, color='9C0006')
f_grn   = F(size=10, bold=True, color='006100')
f_red_b = F(size=12, bold=True, color='9C0006')
f_grn_b = F(size=12, bold=True, color='006100')
f_blue  = F(size=11, bold=True, color='1F3864')

P = lambda c: PatternFill('solid', fgColor=c)
p_hdr    = P('1F3864'); p_gray   = P('F2F2F2'); p_yellow = P('FFF2CC')
p_blue   = P('D6E4F0'); p_sec    = P('D9E2F3'); p_red    = P('FFC7CE')
p_green  = P('C6EFCE'); p_orange = P('FCE4D6'); p_gold   = P('FFEB9C')

ac = Alignment(horizontal='center', vertical='center')
ar = Alignment(horizontal='right',  vertical='center')
al = Alignment(horizontal='left',   vertical='center')

thin = Side(style='thin', color='B0B0B0')
med  = Side(style='medium', color='1F3864')
B_all = Border(left=thin, right=thin, top=thin, bottom=thin)
B_bot = Border(left=thin, right=thin, top=thin, bottom=med)

YEN     = '#,##0"円"'
YEN_NEG = '#,##0"円";▲#,##0"円"'

def s(cell, font=f_body, fill=None, align=al, brd=B_all, fmt=None):
    cell.font=font; cell.alignment=align; cell.border=brd
    if fill: cell.fill=fill
    if fmt:  cell.number_format=fmt

def section(row, text):
    ws.merge_cells(f'B{row}:H{row}')
    ws.cell(row=row,column=2,value=text).font=f_bold
    b=Border(bottom=med)
    for col in range(2,9):
        ws.cell(row=row,column=col).fill=p_sec
        ws.cell(row=row,column=col).border=b
    ws.cell(row=row,column=2).alignment=al

def cf(rng, red_f, grn_f=None, big=False):
    fr,fg = (f_red_b,f_grn_b) if big else (f_red,f_grn)
    ws.conditional_formatting.add(rng, FormulaRule(formula=[red_f], fill=p_red,   font=fr))
    if grn_f:
        ws.conditional_formatting.add(rng, FormulaRule(formula=[grn_f], fill=p_green, font=fg))

for col,w in {'B':4,'C':14,'D':10,'E':14,'F':14,'G':14,'H':22}.items():
    ws.column_dimensions[col].width=w

# ════════════ タイトル (行1-2) ════════════
ws.merge_cells('B1:H1')
c=ws.cell(row=1,column=2,value='送別会 会費収支報告書')
c.font=f_title; c.alignment=ac
ws.row_dimensions[1].height=30
ws.cell(row=2,column=2,value='開催日:').font=f_bold
ws.cell(row=2,column=3,value='2026年4月17日').font=f_body
ws.cell(row=2,column=5,value='参加人数:').font=f_bold
ws.cell(row=2,column=6,value='16名（うち招待3名）').font=f_body
ws.row_dimensions[3].height=6

# ════════════ 参加者一覧 (行4-22) ════════════
section(4,'■ 参加者一覧')
for ci,t in [(2,'No.'),(3,'氏名'),(4,'区分'),(5,'会費'),(6,'ご支援金'),(7,'負担額合計'),(8,'備考')]:
    s(ws.cell(row=5,column=ci,value=t), font=f_hdr, fill=p_hdr, align=ac)

members=[
    ('渡辺','上司',8000,12000,''),('中村','上司',8000,12000,''),
    ('福岡','上司',8000,12000,''),('板倉','上司',8000,12000,''),
    ('幸田','一般',8000,0,''),('宮元','一般',8000,0,''),
    ('武田','一般',8000,0,''),('北野','一般',8000,0,''),
    ('ソラウィット','一般',8000,0,''),('スメート','一般',8000,0,''),
    ('エドセル','一般',8000,0,''),('マール','一般',8000,0,''),
    ('ジャッカパン','一般',8000,0,''),
    ('史','招待',8000,0,'招待対象者'),('大川','招待',8000,0,'招待対象者'),
    ('パユット','招待',8000,0,'招待対象者'),
]
for i,(name,role,fee,sup,note) in enumerate(members):
    r=6+i; rf=p_gray if i%2==1 else None
    cc='1F3864' if role=='上司' else ('C00000' if role=='招待' else '000000')
    ws.cell(row=r,column=2,value=i+1);         s(ws.cell(row=r,column=2), align=ac, fill=rf)
    ws.cell(row=r,column=3,value=name);        s(ws.cell(row=r,column=3), fill=rf)
    ws.cell(row=r,column=4,value=role);        s(ws.cell(row=r,column=4), font=F(size=10,bold=True,color=cc), align=ac, fill=rf)
    ws.cell(row=r,column=5,value=fee);         s(ws.cell(row=r,column=5), align=ar, fill=rf, fmt=YEN)
    ws.cell(row=r,column=6,value=sup);         s(ws.cell(row=r,column=6), align=ar, fill=rf, fmt=YEN)
    ws.cell(row=r,column=7,value=0 if role=='招待' else f'=E{r}+F{r}')
    s(ws.cell(row=r,column=7), align=ar, fill=rf, fmt=YEN)
    ws.cell(row=r,column=8,value=note);        s(ws.cell(row=r,column=8), fill=rf)

# 小計(行22)
rs=22
ws.merge_cells(f'B{rs}:D{rs}')
s(ws.cell(row=rs,column=2,value='小計'), font=f_bold, fill=p_blue, align=ac, brd=B_bot)
for col in [3,4]:
    ws.cell(row=rs,column=col).fill=p_blue; ws.cell(row=rs,column=col).border=B_bot
for col,formula in [(5,'=SUM(E6:E21)'),(6,'=SUM(F6:F21)'),(7,'=SUM(G6:G21)')]:
    ws.cell(row=rs,column=col,value=formula)
    s(ws.cell(row=rs,column=col), font=f_bold, fill=p_blue, align=ar, brd=B_bot, fmt=YEN)
ws.cell(row=rs,column=8).fill=p_blue; ws.cell(row=rs,column=8).border=B_bot
ws.row_dimensions[23].height=6

# ════════════ 固定支出 (行24-29) ════════════
section(24,'■ 固定支出（記念品・交通費等）')
for ci,t in [(2,'No.'),(3,'項目'),(4,''),(5,''),(6,'金額'),(7,''),(8,'備考')]:
    s(ws.cell(row=25,column=ci,value=t), font=f_hdr, fill=p_hdr, align=ac)

for r,name,amt,note,gray in [
    (26,'記念品',15806,'7,853+7,953',False),
    (27,'高速代',580,'',True),
    (28,'交通費',3000,'',False),
]:
    rf=p_gray if gray else None
    ws.cell(row=r,column=2,value=r-25); s(ws.cell(row=r,column=2), align=ac, fill=rf)
    ws.merge_cells(f'C{r}:E{r}')
    ws.cell(row=r,column=3,value=name); s(ws.cell(row=r,column=3), fill=rf)
    for col in [4,5]:
        ws.cell(row=r,column=col).border=B_all
        if rf: ws.cell(row=r,column=col).fill=rf
    ws.merge_cells(f'F{r}:G{r}')
    ws.cell(row=r,column=6,value=amt); s(ws.cell(row=r,column=6), align=ar, fill=rf, fmt=YEN)
    ws.cell(row=r,column=7).border=B_all
    if rf: ws.cell(row=r,column=7).fill=rf
    ws.cell(row=r,column=8,value=note); s(ws.cell(row=r,column=8), fill=rf)

r=29
ws.merge_cells(f'B{r}:E{r}')
s(ws.cell(row=r,column=2,value='固定支出合計'), font=f_bold, fill=p_blue, align=ac, brd=B_bot)
for col in [3,4,5]:
    ws.cell(row=r,column=col).fill=p_blue; ws.cell(row=r,column=col).border=B_bot
ws.merge_cells(f'F{r}:G{r}')
ws.cell(row=r,column=6,value='=SUM(F26:F28)')
s(ws.cell(row=r,column=6), font=f_bold, fill=p_blue, align=ar, brd=B_bot, fmt=YEN)
for col in [7,8]:
    ws.cell(row=r,column=col).fill=p_blue; ws.cell(row=r,column=col).border=B_bot
ws.row_dimensions[30].height=6

# ════════════ 食事・お酒 予算管理 (行31-43) ════════════
section(31,'■ 食事・お酒 予算管理')
for ci,t in [(2,''),(3,'項目'),(4,''),(5,'予算'),(6,'実績'),(7,'残額'),(8,'状態')]:
    s(ws.cell(row=32,column=ci,value=t), font=f_hdr, fill=p_hdr, align=ac)

# 飲食に使える予算（自動計算）行33
r=33
ws.merge_cells(f'B{r}:D{r}')
s(ws.cell(row=r,column=2,value='飲食に使える予算'), font=f_bold, fill=p_orange, align=al)
ws.cell(row=r,column=4).fill=p_orange; ws.cell(row=r,column=4).border=B_all
# 予算列: G22-F29 で自動計算
ws.cell(row=r,column=5,value='=G22-F29')
s(ws.cell(row=r,column=5), font=f_sum, fill=p_orange, align=ar, fmt=YEN)
ws.cell(row=r,column=6).fill=p_orange; ws.cell(row=r,column=6).border=B_all
ws.cell(row=r,column=7).fill=p_orange; ws.cell(row=r,column=7).border=B_all
ws.cell(row=r,column=8,value='← 収入合計 − 固定支出（自動計算）')
s(ws.cell(row=r,column=8), font=f_sub, fill=p_orange, align=al)
ws.row_dimensions[34].height=4

# 食事代 (行35)
r=35
ws.cell(row=r,column=2,value=1); s(ws.cell(row=r,column=2), align=ac)
ws.merge_cells(f'C{r}:D{r}')
ws.cell(row=r,column=3,value='食事代'); s(ws.cell(row=r,column=3))
s(ws.cell(row=r,column=5), fill=p_yellow, align=ar, fmt=YEN)  # 予算入力
s(ws.cell(row=r,column=6), fill=p_yellow, align=ar, fmt=YEN)  # 実績入力
# 残額: 予算と実績どちらか入っていれば表示
ws.cell(row=r,column=7,value=f'=IF(AND(E{r}="",F{r}=""),"",IF(E{r}="",0,E{r})-IF(F{r}="",0,F{r}))')
s(ws.cell(row=r,column=7), align=ar, fmt=YEN_NEG)
ws.cell(row=r,column=8,value=f'=IF(AND(E{r}="",F{r}=""),"←入力してください",IF(G{r}<0,"⚠ 赤字注意！","✓ OK"))')
s(ws.cell(row=r,column=8), font=f_bold, align=ac)

# お酒代 (行36)
r=36
ws.cell(row=r,column=2,value=2); s(ws.cell(row=r,column=2), align=ac, fill=p_gray)
ws.merge_cells(f'C{r}:D{r}')
ws.cell(row=r,column=3,value='お酒代'); s(ws.cell(row=r,column=3), fill=p_gray)
s(ws.cell(row=r,column=5), fill=p_yellow, align=ar, fmt=YEN)
s(ws.cell(row=r,column=6), fill=p_yellow, align=ar, fmt=YEN)
ws.cell(row=r,column=7,value=f'=IF(AND(E{r}="",F{r}=""),"",IF(E{r}="",0,E{r})-IF(F{r}="",0,F{r}))')
s(ws.cell(row=r,column=7), fill=p_gray, align=ar, fmt=YEN_NEG)
ws.cell(row=r,column=8,value=f'=IF(AND(E{r}="",F{r}=""),"←入力してください",IF(G{r}<0,"⚠ 赤字注意！","✓ OK"))')
s(ws.cell(row=r,column=8), font=f_bold, fill=p_gray, align=ac)

cf('G35:H35','$G$35<0','AND($G$35>=0,$G$35<>"")')
cf('G36:H36','$G$36<0','AND($G$36>=0,$G$36<>"")')

# 合計 (行37)
r=37
ws.merge_cells(f'B{r}:D{r}')
s(ws.cell(row=r,column=2,value='合計'), font=f_bold, fill=p_blue, align=ac, brd=B_bot)
ws.cell(row=r,column=4).fill=p_blue; ws.cell(row=r,column=4).border=B_bot
# 予算合計
ws.cell(row=r,column=5,value='=IF(AND(E35="",E36=""),"",IF(E35="",0,E35)+IF(E36="",0,E36))')
s(ws.cell(row=r,column=5), font=f_bold, fill=p_blue, align=ar, brd=B_bot, fmt=YEN)
# 実績合計（バグ修正: F35+F36 を正しく合計）
ws.cell(row=r,column=6,value='=IF(AND(F35="",F36=""),"",IF(F35="",0,F35)+IF(F36="",0,F36))')
s(ws.cell(row=r,column=6), font=f_bold, fill=p_blue, align=ar, brd=B_bot, fmt=YEN)
# 残額合計
ws.cell(row=r,column=7,value='=IF(AND(G35="",G36=""),"",IF(G35="",0,G35)+IF(G36="",0,G36))')
s(ws.cell(row=r,column=7), font=f_bold, fill=p_blue, align=ar, brd=B_bot, fmt=YEN_NEG)
ws.cell(row=r,column=8).fill=p_blue; ws.cell(row=r,column=8).border=B_bot
ws.row_dimensions[38].height=6

# ── あと注文できる金額 (行39) ──
# 飲食予算(E33) − 実績合計。実績が未入力なら予算合計で計算
r=39
ws.merge_cells(f'B{r}:E{r}')
s(ws.cell(row=r,column=2,value='あと注文できる金額（飲食予算 − 実績合計）'),
  font=F(size=12,bold=True), align=al, brd=B_bot)
ws.merge_cells(f'F{r}:G{r}')
# 実績合計(F37)が入力済→実績ベース、未入力なら予算合計(E37)ベース、それも未入力なら飲食予算全額
ws.cell(row=r,column=6,
    value='=E33-IF(F37<>"",F37,IF(E37<>"",E37,0))')
s(ws.cell(row=r,column=6), font=f_big, align=ar, brd=B_bot, fmt=YEN_NEG)
ws.cell(row=r,column=8,value='=IF(F39<0,"⚠ これ以上注文すると赤字！","✓ まだ余裕あり")')
s(ws.cell(row=r,column=8), font=f_bold, align=ac, brd=B_bot)
cf('F39:G39','$F$39<0','$F$39>=0', big=True)
cf('H39','$F$39<0','$F$39>=0')
ws.row_dimensions[40].height=6

# ════════════ 収支サマリー (行41-51) ════════════
section(41,'■ 収支サマリー（全体）')

# 収入
r=42
ws.merge_cells(f'B{r}:E{r}')
s(ws.cell(row=r,column=2,value='収入合計（負担額合計）'), font=f_bold, align=al)
ws.merge_cells(f'F{r}:G{r}')
ws.cell(row=r,column=6,value='=G22'); s(ws.cell(row=r,column=6), font=f_sum, align=ar, fmt=YEN)
ws.cell(row=r,column=8).border=B_all

# 固定支出
r=43
ws.merge_cells(f'B{r}:E{r}')
s(ws.cell(row=r,column=2,value='固定支出（記念品・高速代・交通費）'), font=f_bold, fill=p_gray, align=al)
ws.merge_cells(f'F{r}:G{r}')
ws.cell(row=r,column=6,value='=F29'); s(ws.cell(row=r,column=6), font=f_sum, fill=p_gray, align=ar, fmt=YEN)
ws.cell(row=r,column=8).fill=p_gray; ws.cell(row=r,column=8).border=B_all

# 飲食費
r=44
ws.merge_cells(f'B{r}:E{r}')
s(ws.cell(row=r,column=2,value='飲食費（実績入力済なら実績、未入力なら計画値16名×8,000円）'), font=f_bold, align=al)
ws.merge_cells(f'F{r}:G{r}')
# 実績合計F37が入力済ならそれを使用、未入力なら 16名×8000=E22
ws.cell(row=r,column=6,value='=IF(F37<>"",F37,E22)')
s(ws.cell(row=r,column=6), font=f_sum, align=ar, fmt=YEN)
ws.cell(row=r,column=8,value='← 実績入力で自動切替').font=f_sub
ws.cell(row=r,column=8).border=B_all

# 差引残高
r=45
ws.merge_cells(f'B{r}:E{r}')
s(ws.cell(row=r,column=2,value='差引残高'), font=F(size=12,bold=True), align=al, brd=B_bot)
ws.merge_cells(f'F{r}:G{r}')
ws.cell(row=r,column=6,value='=F42-F43-F44')
s(ws.cell(row=r,column=6), font=f_big, align=ar, brd=B_bot, fmt=YEN_NEG)
ws.cell(row=r,column=8,value='=IF(F45<0,"⚠ 赤字！","✓ 黒字")')
s(ws.cell(row=r,column=8), font=f_bold, align=ac, brd=B_bot)
cf('F45:G45','$F$45<0','$F$45>=0', big=True)
cf('H45','$F$45<0','$F$45>=0')

ws.row_dimensions[46].height=6

# 1人あたり負担額
r=47
ws.merge_cells(f'B{r}:E{r}')
s(ws.cell(row=r,column=2,value='一般参加者の実質負担額（8,000円 − 返金額）'), font=f_bold, align=al)
ws.merge_cells(f'F{r}:G{r}')
ws.cell(row=r,column=6,value='=8000-IF(F45>0,ROUND(F45/9,0),0)')
s(ws.cell(row=r,column=6), font=f_blue, align=ar, fmt=YEN)
ws.cell(row=r,column=8).border=B_all

ws.row_dimensions[48].height=6

# ── 返金額（黒字の場合）行49-50 ──
r=49
ws.merge_cells(f'B{r}:H{r}')
s(ws.cell(row=r,column=2,value='▼ 余剰金が発生した場合の返金'), font=f_bold, fill=p_gold, align=al,
  brd=Border(bottom=Side(style='medium', color='C09000')))
for col in range(3,9):
    ws.cell(row=r,column=col).fill=p_gold
    ws.cell(row=r,column=col).border=Border(bottom=Side(style='medium', color='C09000'))

r=50
ws.merge_cells(f'B{r}:E{r}')
s(ws.cell(row=r,column=2,value='1人あたり返金額（上司を除く一般参加者9名で按分）'), font=f_bold, fill=p_gold, align=al)
ws.merge_cells(f'F{r}:G{r}')
# 一般参加者: 16名 - 招待3名 - 上司4名 = 9名
ws.cell(row=r,column=6,value='=IF(F45<=0,"返金なし",ROUND(F45/9,0)&"円")')
s(ws.cell(row=r,column=6), font=F(size=12,bold=True,color='7B4900'), fill=p_gold, align=ar)
ws.cell(row=r,column=8,value='対象：一般参加者9名').font=f_sub
ws.cell(row=r,column=8).fill=p_gold; ws.cell(row=r,column=8).border=B_all

# 黒字時に返金セルをハイライト
ws.conditional_formatting.add(f'F{r}:G{r}',
    FormulaRule(formula=['$F$45>0'], fill=p_gold,
                font=F(size=12,bold=True,color='7B4900')))

# 注記
ws.row_dimensions[51].height=6
ws.merge_cells('B52:H52')
ws.cell(row=52,column=2,value='※ 黄色セルに食事代・お酒代の「予算」「実績」を入力 → 残額・警告・返金額が自動更新されます').font=f_note
ws.merge_cells('B53:H53')
ws.cell(row=53,column=2,value='※ 招待対象者（史・大川・パユット）は実質負担0円').font=f_note

# 印刷設定
ws.page_setup.orientation='landscape'
ws.page_setup.paperSize=ws.PAPERSIZE_A4
ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
ws.sheet_properties.pageSetUpPr.fitToPage=True
ws.print_area='B1:H53'

output='C:/Users/SEIGI-N13/Desktop/260417送別会 会費_改善版.xlsx'
wb.save(output)
print('完了')

# 会費収支管理テンプレート v3 再構築
#  - 参加者 最大40名
#  - 開催店舗/住所/URL 欄（設定シート）
#  - 参加者ごとの状況：アルコール/食事制限/送迎（ドロップダウン＋集計）
#  - 固定費に「送迎バス代」を追加
#  - 役職に「技師」「主任」を追加（社員と同じ一律返金グループ・ご支援金なし）
import sys
sys.stdout.reconfigure(encoding="utf-8")
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

SRC = "会費収支管理テンプレート_20260714A.xlsx"
DST = "会費収支管理テンプレート_20260714C.xlsx"

RANKS = ["社員", "技師", "主任", "係長", "課長", "次長", "部長", "専務", "常務", "顧問", "社長"]
FLAT_GROUP = ["社員", "技師", "主任", "係長"]   # 一律返金グループ（ご支援金なし）。設定シート B14:B17
DEFAULT_FEE = 8000
DEFAULT_SUPPORT = 12000
N = 40  # 参加者最大人数

wb = load_workbook(SRC)
old_set = wb["設定"]
old_mgmt = wb["収支管理"]

# ---- 書式プロキシを既存セルから採取（同一wb内なので安全）----
S = {
    "title": copy(old_mgmt["B1"]._style),      # 濃紺タイトル白字
    "section": copy(old_mgmt["B4"]._style),    # ■見出し 濃紺帯
    "header": copy(old_mgmt["C5"]._style),     # 表ヘッダー 青帯白字中央
    "input": copy(old_mgmt["C6"]._style),      # 入力 黄 左
    "input_c": copy(old_mgmt["D6"]._style),    # 入力 黄 中央
    "auto": copy(old_mgmt["E6"]._style),       # 自動計算 水色 右
    "note_s": copy(old_mgmt["H6"]._style),     # 備考 黄 小
    "subtotal": copy(old_mgmt["B36"]._style),  # 小計 青帯
    "subtotal_r": copy(old_mgmt["E36"]._style),# 小計 青帯 右
    "cnt": copy(old_mgmt["B37"]._style),       # 人数サマリー
    "sub_head": copy(old_mgmt["B53"]._style),  # 【居酒屋】青見出し
    "note": copy(old_mgmt["B52"]._style),      # 注記
    "num": copy(old_mgmt["B41"]._style),       # No.セル
    "item": copy(old_mgmt["C41"]._style),      # 項目名 黄
    "budget": copy(old_mgmt["E41"]._style),    # 予算 黄 右
    "actual": copy(old_mgmt["F41"]._style),    # 実績 黄 右
    "diff": copy(old_mgmt["G41"]._style),      # 差額 自動
    "smry_l": copy(old_mgmt["B82"]._style),    # サマリーラベル 薄緑
    "smry_v": copy(old_mgmt["F82"]._style),    # サマリー値 薄緑 右
    "bal_l": copy(old_mgmt["B85"]._style),     # 差引残高 水色太
    "bal_v": copy(old_mgmt["F85"]._style),     # 差引残高値 水色太
    "warn": copy(old_mgmt["G85"]._style),      # 警告
}
FMT_YEN = "#,##0"

# 設定シートの入力欄スタイル（v1設定シートの正しいセルから採取）
Sset = {
    "title": copy(old_set["B1"]._style),
    "section": copy(old_set["B9"]._style),   # ■見出し
    "label": copy(old_set["B4"]._style),     # 項目ラベル
    "input": copy(old_set["C4"]._style),     # 入力欄（黄）
    "note": copy(old_set["B20"]._style),     # 注記
    "header": copy(old_mgmt["C5"]._style),   # 表ヘッダー（青）
    "auto": copy(old_mgmt["E6"]._style),     # 自動計算（水色）
}

# ---- 旧シートを削除して同位置に作り直す ----
idx_set = wb.sheetnames.index("設定")
idx_mgmt = wb.sheetnames.index("収支管理")
del wb["設定"]
del wb["収支管理"]
ws_set = wb.create_sheet("設定", idx_set)
ws = wb.create_sheet("収支管理", idx_mgmt)


def put(sheet, addr, value, style, fmt=None):
    c = sheet[addr]
    c.value = value
    c._style = copy(style)
    if fmt:
        c.number_format = fmt
    return c


def band(sheet, rng, value, style, fmt=None):
    """結合セル：範囲全体に同じ塗り、左上に値"""
    from openpyxl.utils import range_boundaries, get_column_letter
    minc, minr, maxc, maxr = range_boundaries(rng)
    for r in range(minr, maxr + 1):
        for cc in range(minc, maxc + 1):
            cell = sheet.cell(row=r, column=cc)
            cell._style = copy(style)
    anchor = f"{get_column_letter(minc)}{minr}"
    sheet[anchor].value = value
    if fmt:
        sheet[anchor].number_format = fmt
    sheet.merge_cells(rng)


# =====================================================================
# 設定シート
# =====================================================================
ws_set.column_dimensions["A"].width = 2
ws_set.column_dimensions["B"].width = 32
ws_set.column_dimensions["C"].width = 22
ws_set.column_dimensions["D"].width = 16
ws_set.column_dimensions["E"].width = 16

band(ws_set, "B1:E1", "⚙️  イベント基本設定", Sset["title"])

band(ws_set, "B3:E3", "■ イベント情報", Sset["section"])
info = [
    ("イベント種別（必須）★", "バーベキュー"),
    ("イベント名", None),
    ("開催日", None),
    ("幹事名", None),
    ("開催店舗（店名）", None),
    ("開催場所　住所", None),
    ("開催店舗　URL", None),
]
for i, (lab, val) in enumerate(info):
    r = 4 + i
    put(ws_set, f"B{r}", lab, Sset["label"])
    band(ws_set, f"C{r}:E{r}", val, Sset["input"])

# イベント種別ドロップダウン
dv_kind = DataValidation(type="list", formula1='"居酒屋,バーベキュー,仕入れ対応（接待）"', allow_blank=True)
ws_set.add_data_validation(dv_kind)
dv_kind.add("C4:E4")

# 会費・ご支援金 役職テーブル
band(ws_set, "B12:E12", "■ 会費・ご支援金設定（役職ごとに変更可）", Sset["section"])
for col, text in [("B", "役職"), ("C", "会費（円）"), ("D", "ご支援金（円）"), ("E", "実質負担（自動）")]:
    put(ws_set, f"{col}13", text, Sset["header"])
for i, rank in enumerate(RANKS):
    r = 14 + i
    put(ws_set, f"B{r}", rank, Sset["label"])
    put(ws_set, f"C{r}", DEFAULT_FEE, Sset["input"], FMT_YEN)
    put(ws_set, f"D{r}", 0 if rank in FLAT_GROUP else DEFAULT_SUPPORT, Sset["input"], FMT_YEN)
    refund = "収支管理!$G$103" if rank in FLAT_GROUP else "収支管理!$G$105"   # 実績（確定）返金を実質負担に反映
    put(ws_set, f"E{r}", f"=C{r}+D{r}-{refund}", Sset["auto"], FMT_YEN)
band(ws_set, "B25:E25", "※ 金額は自由に編集できます（ご支援金なしの役職は 0）。実質負担は返金額を差し引いた自動計算です", Sset["note"])

band(ws_set, "B27:E27", "■ 返金設定", Sset["section"])
put(ws_set, "B28", "一般（社員・技師・主任・係長）　1人あたり返金額（円）", Sset["label"])
put(ws_set, "C28", 3000, Sset["input"], FMT_YEN)
band(ws_set, "D28:E28", "← 余剰金からこの金額を一般（社員・技師・主任・係長）に返金、残額を役職者で按分", Sset["note"])

band(ws_set, "B30:E30", "■ 返金ロジック（参考）", Sset["section"])
band(ws_set, "B31:E31", "①一般（社員・技師・主任・係長）1人あたり：設定額（余剰金が足りない場合は自動で減額）  ②役職者（課長以上）1人あたり：（余剰金－一般返金合計）÷ 役職者人数（端数切捨て）", Sset["note"])

# =====================================================================
# 収支管理シート
# =====================================================================
widths = {"A": 1.5, "B": 6, "C": 22, "D": 11, "E": 14, "F": 14, "G": 14,
          "H": 11, "I": 13, "J": 11, "K": 26}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# 行高（テーブル部）
ws.row_dimensions[1].height = 32.1

# タイトル
band(ws, "B1:K1", '=設定!C5&"　会費収支報告書"', S["title"])

# 開催情報（行2-3）
put(ws, "B2", "開催日：", S["cnt"])
band(ws, "C2:D2", "=設定!C6", S["cnt"])
put(ws, "E2", "種別：", S["cnt"])
band(ws, "F2:G2", "=設定!C4", S["cnt"])
put(ws, "H2", "幹事：", S["cnt"])
band(ws, "I2:K2", "=設定!C7", S["cnt"])
put(ws, "B3", "会場：", S["cnt"])
band(ws, "C3:D3", "=設定!C8", S["cnt"])
put(ws, "E3", "住所：", S["cnt"])
band(ws, "F3:G3", "=設定!C9", S["cnt"])
put(ws, "H3", "URL：", S["cnt"])
band(ws, "I3:K3", "=設定!C10", S["cnt"])

# 参加者一覧見出し（行5）
band(ws, "B5:K5", "■ 参加者一覧（最大40名）　D列の区分・H〜J列の状況はドロップダウンで選択してください", S["section"])
# ヘッダー（行6）
heads = ["No.", "氏名", "区分", "会費", "ご支援金", "負担額合計", "アルコール", "食事制限", "送迎", "備考"]
for i, h in enumerate(heads):
    put(ws, f"{chr(ord('B')+i)}6", h, S["header"])

# 参加者行（7〜46）
first, last = 7, 46
for r in range(first, last + 1):
    ws.row_dimensions[r].height = 17.1
    put(ws, f"B{r}", r - first + 1, S["num"])
    put(ws, f"C{r}", None, S["input"])       # 氏名
    put(ws, f"D{r}", None, S["input_c"])     # 区分
    put(ws, f"E{r}",
        f'=IF(D{r}="","",IF(OR(D{r}="招待",D{r}="欠席"),0,'
        f'IFERROR(INDEX(設定!$C$14:$C$24,MATCH(D{r},設定!$B$14:$B$24,0)),0)))',
        S["auto"], FMT_YEN)                  # 会費
    put(ws, f"F{r}",
        f'=IF(OR(D{r}="",D{r}="招待",D{r}="欠席"),0,'
        f'IFERROR(INDEX(設定!$D$14:$D$24,MATCH(D{r},設定!$B$14:$B$24,0)),0))',
        S["auto"], FMT_YEN)                  # ご支援金
    put(ws, f"G{r}", f'=IF(D{r}="","",E{r}+F{r})', S["auto"], FMT_YEN)  # 合計
    put(ws, f"H{r}", None, S["input_c"])     # アルコール
    put(ws, f"I{r}", None, S["input_c"])     # 食事制限
    put(ws, f"J{r}", None, S["input_c"])     # 送迎
    put(ws, f"K{r}", None, S["note_s"])      # 備考

# カウント式（共通）
CNT_YAKU = "SUMPRODUCT(COUNTIF(D7:D46,設定!$B$18:$B$24))"   # 役職者＝課長以上（設定B18:B24）
CNT_SHAIN = "SUMPRODUCT(COUNTIF(D7:D46,設定!$B$14:$B$17))"  # 一般＝社員・技師・主任・係長（設定B14:B17）
PAX = f"({CNT_SHAIN}+{CNT_YAKU})"  # 会費負担人数（招待・欠席を除く）＝返金・1人当たり実績の分母
BUDGET_PAX = f'({CNT_SHAIN}+{CNT_YAKU}+COUNTIF(D7:D46,"招待"))'  # 予算の頭数＝出席者（招待含む・欠席除く）。招待者も社員と同じ単価で費用計上

# 小計（行47）
band(ws, "B47:D47", "小計", S["subtotal"])
put(ws, "E47", "=SUM(E7:E46)", S["subtotal_r"], FMT_YEN)
put(ws, "F47", "=SUM(F7:F46)", S["subtotal_r"], FMT_YEN)
put(ws, "G47", "=SUM(G7:G46)", S["subtotal_r"], FMT_YEN)
for col in ["H", "I", "J", "K"]:
    put(ws, f"{col}47", None, S["subtotal"])

# 人数サマリー（行48）
band(ws, "B48:K48",
     f'={CNT_SHAIN}&"名（一般）　"&{CNT_YAKU}&"名（役職者）　"'
     '&COUNTIF(D7:D46,"招待")&"名（招待）　"&COUNTIF(D7:D46,"欠席")&"名（欠席）　"'
     '&"合計 "&COUNTA(C7:C46)&"名"',
     S["cnt"])

# 状況サマリー（行49）
band(ws, "B49:K49",
     '="🍺 アルコール：あり "&COUNTIF(H7:H46,"あり")&"名／なし "&COUNTIF(H7:H46,"なし")&"名　　"'
     '&"🍽 食事制限あり "&COUNTIF(I7:I46,"あり")&"名（内容は備考欄を確認）　　"'
     '&"🚐 送迎：あり "&COUNTIF(J7:J46,"あり")&"名／なし "&COUNTIF(J7:J46,"なし")&"名"',
     S["smry_l"])

# ---- 固定費（行51〜62）----
# 列: E=1人当たり予算 / F=総額予算 / G=予算合計(自動:総額優先) / H=実績 / I=差額 / J=1人当たり実績 / K=備考
band(ws, "B51:K51", "■ 固定費（確定費用）", S["section"])
for col, text in [("B", "No."), ("C", "項目名"), ("E", "1人当たり予算"), ("F", "総額予算"), ("G", "予算合計"), ("H", "実績（円）"), ("I", "差額"), ("J", "1人当たり実績"), ("K", "備考")]:
    put(ws, f"{col}52", text, S["header"])
put(ws, "D52", None, S["header"])
fixed_items = ["記念品", "送迎バス代", "交通費①", "交通費②", "その他①", "その他②", "その他③", "その他④", "その他⑤"]
fx_first = 53
for i, name in enumerate(fixed_items):
    r = fx_first + i
    put(ws, f"B{r}", i + 1, S["num"])
    band(ws, f"C{r}:D{r}", name, S["item"])
    put(ws, f"E{r}", None, S["budget"], FMT_YEN)                                                       # 1人当たり予算（入力）
    put(ws, f"F{r}", None, S["budget"], FMT_YEN)                                                       # 総額予算（入力）
    put(ws, f"G{r}", f'=IF(F{r}<>"",F{r},IF(E{r}<>"",E{r}*{BUDGET_PAX},""))', S["diff"], FMT_YEN)      # 予算合計（総額優先）
    put(ws, f"H{r}", None, S["actual"], FMT_YEN)                                                       # 実績（入力）
    put(ws, f"I{r}", f'=IF(OR(G{r}="",H{r}=""),"",G{r}-H{r})', S["diff"], FMT_YEN)                     # 差額
    put(ws, f"J{r}", f'=IF(OR(H{r}="",{PAX}=0),"",ROUND(H{r}/{PAX},0))', S["diff"], FMT_YEN)           # 1人当たり実績
    put(ws, f"K{r}", None, S["note_s"])                                                                # 備考
fx_last = fx_first + len(fixed_items) - 1  # 61
band(ws, f"B62:D62", "固定費合計", S["subtotal"])
put(ws, "E62", None, S["subtotal"]); put(ws, "F62", None, S["subtotal"])
put(ws, "G62", f"=SUM(G{fx_first}:G{fx_last})", S["subtotal_r"], FMT_YEN)                              # 予算合計
put(ws, "H62", f"=SUM(H{fx_first}:H{fx_last})", S["subtotal_r"], FMT_YEN)                              # 実績合計
put(ws, "I62", f'=IF(SUM(G{fx_first}:G{fx_last})=0,"",SUM(G{fx_first}:G{fx_last})-SUM(H{fx_first}:H{fx_last}))', S["subtotal_r"], FMT_YEN)
put(ws, "J62", f'=IF({PAX}=0,"",ROUND(H62/{PAX},0))', S["subtotal_r"], FMT_YEN)
put(ws, "K62", None, S["subtotal"])

# ---- 変動費（行64〜92）----
band(ws, "B64:K64", "■ 変動費（下記3セクションのうち、使用するものだけ入力してください）", S["section"])
band(ws, "B65:K65", "※ 予算は「1人当たり予算」か「総額予算」のどちらかに入力（総額を入れたらそちらを優先）。1人当たりは招待含む出席者数を掛けます。使わないセクションは空白でOK。", S["note"])

def variable_section(title, start_head, items):
    """見出し行=start_head-1, ヘッダー=start_head, 明細, 実績合計行(tot) を返す
       列: E=1人当たり予算 / F=総額予算 / G=予算合計(自動:総額優先) / H=実績 / I=残額 / J=1人当たり実績 / K=状態"""
    hd = start_head
    band(ws, f"B{hd-1}:K{hd-1}", title, S["sub_head"])
    for col, text in [("B", "No."), ("C", "項目名"), ("E", "1人当たり予算"), ("F", "総額予算"), ("G", "予算合計"), ("H", "実績（円）"), ("I", "残額"), ("J", "1人当たり実績"), ("K", "状態")]:
        put(ws, f"{col}{hd}", text, S["header"])
    put(ws, f"D{hd}", None, S["header"])
    di = hd + 1
    for i, name in enumerate(items):
        r = di + i
        put(ws, f"B{r}", i + 1, S["num"])
        band(ws, f"C{r}:D{r}", name, S["item"])
        put(ws, f"E{r}", None, S["budget"], FMT_YEN)                                                       # 1人当たり予算（入力）
        put(ws, f"F{r}", None, S["budget"], FMT_YEN)                                                       # 総額予算（入力）
        put(ws, f"G{r}", f'=IF(F{r}<>"",F{r},IF(E{r}<>"",E{r}*{BUDGET_PAX},""))', S["diff"], FMT_YEN)      # 予算合計（総額優先）
        put(ws, f"H{r}", None, S["actual"], FMT_YEN)                                                       # 実績（入力）
        put(ws, f"I{r}", f'=IF(OR(G{r}="",H{r}=""),"",G{r}-H{r})', S["diff"], FMT_YEN)                     # 残額
        put(ws, f"J{r}", f'=IF(OR(H{r}="",{PAX}=0),"",ROUND(H{r}/{PAX},0))', S["diff"], FMT_YEN)           # 1人当たり実績
        put(ws, f"K{r}", f'=IF(OR(G{r}="",H{r}=""),"",IF(I{r}<0,"⚠ 予算超過","✓ OK"))', S["diff"])         # 状態
    dl = di + len(items) - 1
    tot = dl + 1
    band(ws, f"B{tot}:D{tot}", title.strip("【】 ") + " 実績合計", S["subtotal"])
    put(ws, f"G{tot}", f"=SUM(G{di}:G{dl})", S["subtotal_r"], FMT_YEN)                        # 予算合計
    put(ws, f"H{tot}", f"=SUM(H{di}:H{dl})", S["subtotal_r"], FMT_YEN)                        # 実績合計
    put(ws, f"J{tot}", f'=IF({PAX}=0,"",ROUND(H{tot}/{PAX},0))', S["subtotal_r"], FMT_YEN)
    for col in ["E", "F", "I", "K"]:
        put(ws, f"{col}{tot}", None, S["subtotal"])
    return tot

izakaya_tot = variable_section("【 居酒屋 】", 67, ["食事代", "お酒代・飲み放題代", "コース追加", "その他"])          # 見出し66,ヘッダー67,68-71,合計72
bbq_tot = variable_section("【 バーベキュー 】", 75, ["食材費（肉・野菜等）", "炭・着火剤・備品代", "会場使用料", "機材レンタル代", "その他"])  # 見出し74,ヘッダー75,76-80,合計81
shiire_tot = variable_section("【 仕入れ対応（接待） 】", 84, ["食材費", "消耗品費（割箸・紙皿等）", "飲み物代", "装飾・準備費", "その他"])   # 見出し83,ヘッダー84,85-89,合計90

var_total_row = 92
band(ws, f"B{var_total_row}:D{var_total_row}", "変動費合計（全セクション）", S["subtotal"])
put(ws, f"G{var_total_row}", f"=G{izakaya_tot}+G{bbq_tot}+G{shiire_tot}", S["subtotal_r"], FMT_YEN)  # 予算合計
put(ws, f"H{var_total_row}", f"=H{izakaya_tot}+H{bbq_tot}+H{shiire_tot}", S["subtotal_r"], FMT_YEN)  # 実績合計
put(ws, f"J{var_total_row}", f'=IF({PAX}=0,"",ROUND(H{var_total_row}/{PAX},0))', S["subtotal_r"], FMT_YEN)
for col in ["E", "F", "I", "K"]:
    put(ws, f"{col}{var_total_row}", None, S["subtotal"])

# ---- 収支サマリー（予算／実績 2列。行94〜99）----
band(ws, "B94:K94", "■ 収支サマリー（予算／実績）", S["section"])
band(ws, "B95:E95", "項目", S["header"]); put(ws, "F95", "予算（見込み）", S["header"]); put(ws, "G95", "実績（確定）", S["header"]); band(ws, "H95:K95", "実績の内訳・注記", S["header"])
band(ws, "B96:E96", "収入合計（会費＋ご支援金）", S["smry_l"]); put(ws, "F96", "=G47", S["smry_v"], FMT_YEN); put(ws, "G96", "=G47", S["smry_v"], FMT_YEN); band(ws, "H96:K96", "会費・ご支援金は確定額（予算・実績とも同じ）", S["warn"])
band(ws, "B97:E97", "固定費合計", S["smry_l"]); put(ws, "F97", "=G62", S["smry_v"], FMT_YEN); put(ws, "G97", "=H62", S["smry_v"], FMT_YEN); band(ws, "H97:K97", None, S["warn"])
band(ws, "B98:E98", "変動費合計", S["smry_l"]); put(ws, "F98", f"=G{var_total_row}", S["smry_v"], FMT_YEN); put(ws, "G98", f"=H{var_total_row}", S["smry_v"], FMT_YEN); band(ws, "H98:K98", None, S["warn"])
band(ws, "B99:E99", "差引残高（余剰金）", S["bal_l"])
put(ws, "F99", '=IF(AND(F97=0,F98=0),"",F96-F97-F98)', S["bal_v"], FMT_YEN)
put(ws, "G99", '=IF(AND(G97=0,G98=0),"",G96-G97-G98)', S["bal_v"], FMT_YEN)
band(ws, "H99:K99", '=IF(G99="","（実績未入力）",IF(G99<0,"⚠ 実績赤字！費用を見直してください","✓ 実績黒字 ─ 下の実績列で返金額を確認"))', S["warn"])

# ---- 返金配分（予算／実績 2列。行101〜108）----
band(ws, "B101:K101", "■ 返金配分（予算／実績）", S["section"])
band(ws, "B102:E102", "項目", S["header"]); put(ws, "F102", "予算（見込み）", S["header"]); put(ws, "G102", "実績（確定）", S["header"]); band(ws, "H102:K102", "実績の内訳・注記", S["header"])
band(ws, "B103:E103", "一般（社員・技師・主任・係長）　1人あたり返金額（自動調整）", S["smry_l"])
put(ws, "F103", f'=IF(OR(F99="",F99<=0,{CNT_SHAIN}=0),0,MIN(設定!$C$28,ROUNDDOWN(F99/{CNT_SHAIN},0)))', S["smry_v"], FMT_YEN)
put(ws, "G103", f'=IF(OR(G99="",G99<=0,{CNT_SHAIN}=0),0,MIN(設定!$C$28,ROUNDDOWN(G99/{CNT_SHAIN},0)))', S["smry_v"], FMT_YEN)
band(ws, "H103:K103", f'=IF(OR(G99="",G99<=0),"実績未確定 または 返金なし","1人 "&TEXT(G103,"#,##0")&"円 × "&{CNT_SHAIN}&"名"&IF(G103<設定!$C$28,"　※余剰金不足のため設定額 "&TEXT(設定!$C$28,"#,##0")&"円 から自動減額",""))', S["warn"])
band(ws, "B104:E104", "一般　返金合計", S["smry_l"]); put(ws, "F104", f"=F103*{CNT_SHAIN}", S["smry_v"], FMT_YEN); put(ws, "G104", f"=G103*{CNT_SHAIN}", S["smry_v"], FMT_YEN); band(ws, "H104:K104", None, S["warn"])
band(ws, "B105:E105", "役職者（課長以上）　1人あたり返金額（残額を按分・切捨て）", S["smry_l"])
put(ws, "F105", f'=IF(OR(F99="",{CNT_YAKU}=0),0,ROUNDDOWN(MAX(F99-F104,0)/{CNT_YAKU},0))', S["smry_v"], FMT_YEN)
put(ws, "G105", f'=IF(OR(G99="",{CNT_YAKU}=0),0,ROUNDDOWN(MAX(G99-G104,0)/{CNT_YAKU},0))', S["smry_v"], FMT_YEN)
band(ws, "H105:K105", f'=IF(OR(G99="",G105=0),"実績未確定 または 返金なし","1人 "&TEXT(G105,"#,##0")&"円 × "&{CNT_YAKU}&"名")', S["warn"])
band(ws, "B106:E106", "一般　実質負担額（税込・社員基準）", S["smry_l"])
put(ws, "F106", f'=IF({CNT_SHAIN}=0,"",設定!$C$14-F103)', S["smry_v"], FMT_YEN)
put(ws, "G106", f'=IF({CNT_SHAIN}=0,"",設定!$C$14-G103)', S["smry_v"], FMT_YEN)
band(ws, "H106:K106", '=TEXT(設定!$C$14,"#,##0")&"円（会費）－実績返金 "&TEXT(G103,"#,##0")&"円"', S["warn"])
band(ws, "B107:E107", "役職者　実質負担額", S["smry_l"]); band(ws, "F107:G107", "設定シート参照", S["smry_v"]); band(ws, "H107:K107", "役職により異なるため、設定シートの「実質負担（自動）」列で確認（実績ベース）", S["warn"])
band(ws, "B108:K108", "※ 「予算」＝1人当たり予算×出席者数（招待含む）の見込み。「実績」＝実際にかかった費用からの確定額。端数は切捨て。招待者は実質負担0円、欠席者は会費0円。", S["note"])

# ---- ドロップダウン ----
dv_kubun = DataValidation(type="list", formula1='"' + ",".join(RANKS) + ',招待,欠席"', allow_blank=True)
dv_kubun.error = "リストから選択してください"; dv_kubun.errorTitle = "入力エラー"
ws.add_data_validation(dv_kubun); dv_kubun.add("D7:D46")

dv_alc = DataValidation(type="list", formula1='"あり,なし"', allow_blank=True)
ws.add_data_validation(dv_alc); dv_alc.add("H7:H46")
dv_meal = DataValidation(type="list", formula1='"あり,なし"', allow_blank=True)
ws.add_data_validation(dv_meal); dv_meal.add("I7:I46")
dv_ride = DataValidation(type="list", formula1='"あり,なし"', allow_blank=True)
ws.add_data_validation(dv_ride); dv_ride.add("J7:J46")

# 使い方シート更新
ws_guide = wb["使い方"]
ws_guide["C4"] = (
    "① イベント種別をドロップダウンで選択\n"
    "   → 居酒屋 / バーベキュー / 仕入れ対応（接待）\n"
    "② イベント名・開催日・幹事名・開催店舗・住所・URLを入力\n"
    "③ 役職ごとの会費・ご支援金を確認・変更（初期値：会費 8,000円 / ご支援金 係長以上 12,000円）\n"
    "④ 一般（社員・技師・主任・係長）への返金額を設定（初期値：3,000円）"
)
ws_guide["C7"] = (
    "① C列（氏名）に名前を入力（最大40名）\n"
    "② D列（区分）を役職のドロップダウンで選択\n"
    "   社員 / 技師 / 主任 / 係長 / 課長 / 次長 / 部長 / 専務 / 常務 / 顧問 / 社長 / 招待 / 欠席\n"
    "   ※社員・技師・主任・係長は「一般」（ご支援金なし・一律返金）、課長以上は「役職者」（ご支援金あり・按分返金）\n"
    "③ E（会費）・F（支援金）は役職に応じて自動入力\n"
    "④ H〜J列で各人の状況を選択：アルコール（あり/なし）・食事制限（あり/なし）・送迎（あり/なし）\n"
    "   ※食事制限「あり」の場合は K列（備考）に内容を記入\n"
    "⑤ 状況の集計は参加者一覧の下（状況サマリー行）に自動表示"
)
ws_guide["C16"] = (
    "差引残高・返金金額はすべて自動計算されます\n"
    "✓ 黒字 → 一般（社員・技師・主任・係長）へ返金（余剰金が足りなければ自動で減額）、残額を役職者（課長以上）で按分\n"
    "⚠ 赤字 → 会費や費用の見直しが必要です\n"
    "端数は切捨て（2〜3円程度の予備費が自動保留されます）"
)

wb.save(DST)
print("保存完了:", DST)

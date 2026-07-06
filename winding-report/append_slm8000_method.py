# -*- coding: utf-8 -*-
"""
AISIN軌跡計算ファイルに SLM8000方式 シートを追加／更新するスクリプト（v3）

【目的】
  既存 AISIN方式 軌跡計算（LX/LZシート・20230825系シート）を温存しつつ、
  SLM8000方式 の軌道計算を左右（S軸・U軸）2テーブル＋2グラフで実装。
  各グラフは指令値(Cmd)・実測値(Act)・製品輪郭を 1 枚に統合し色分けで比較する。

【v3 の変更点（2026-04-17 2回目）】
  - LZ軸の単位表記を [°] → [mm] に変更
  - 軸タイトル "LX 方向" → "X 方向"、"LZ 方向" → "Z 方向"
  - 右テーブルから LZ 重複列を削除（LZはグラフ系列として左表から直接参照）
  - グラフを大きく（22cm × 17cm）。凡例を右側に移動して軸タイトルとの被り解消
  - 非入力セル（STEP・時刻・計算列）を明確に灰色化し、入力セル（水色/橙色）と区別
  - 列幅を信号名が切れないよう調整し、wrap_text で対応
  - 再実行時は既存の入力データ（C/D/E/F/N/O 列）を保持したまま再生成

【ユーザー合意事項】
  - 2パターン = AISIN方式 と SLM8000方式（既存温存）
  - ワーク寸法 = AISIN の L1・R 幾何寸法 ＋ 20230825 シート製品図形座標
  - データ入力 = コピペ方式
  - 軸マッピング：θ → LX（S軸/U軸・振り軸）、Z → LZ（上下軸）
  - 左右軸：左=S軸 / 右=U軸 / LZ軸は左右共通（1本）

【運用想定】
  1. Omron Sysmac Studio データトレース実行
  2. CSV 出力：
     - MC_Axis_S.Cmd.Pos / .Act.Pos   [°]  → 左テーブル C / D
     - MC_Axis_LZ.Cmd.Pos / .Act.Pos  [mm] → 左テーブル E / F（LZは左右共通なのでここだけ）
     - MC_Axis_U.Cmd.Pos / .Act.Pos   [°]  → 右テーブル N / O
  3. 左右グラフに指令・実測・製品輪郭の 3 系列が色分けで自動更新
"""

import sys
import os
import string
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

# ════════════════════════════════════════
# 設定
# ════════════════════════════════════════
DESKTOP_DIR = r"C:\Users\SEIGI-N13\Desktop"
BASE_NAME   = "2976 LXLZ軌跡計算式"
EXT         = ".xlsx"
SHEET_NEW   = "SLM8000方式_軌道計算"

# ファイル命名規則：{ベース名}_{YYYYMMDD}{A-Z}.xlsx
#   A=その日の初回更新、B=2回目、以降C, D…と増加。日付が変わったらAリセット
def _today():
    return datetime.now().strftime('%Y%m%d')

def find_latest_input():
    """最新の入力ファイルを探す。今日の最新ラベル → レガシー _NEW → オリジナル の順"""
    td = _today()
    for letter in reversed(string.ascii_uppercase):
        p = os.path.join(DESKTOP_DIR, f"{BASE_NAME}_{td}{letter}{EXT}")
        if os.path.exists(p):
            return p
    legacy = os.path.join(DESKTOP_DIR, f"{BASE_NAME}_20230525_NEW.xlsx")
    if os.path.exists(legacy):
        return legacy
    return os.path.join(DESKTOP_DIR, f"{BASE_NAME}_20230525.xlsx")

def next_output_path():
    """今日の次の空きラベルで出力パスを生成"""
    td = _today()
    for letter in string.ascii_uppercase:
        p = os.path.join(DESKTOP_DIR, f"{BASE_NAME}_{td}{letter}{EXT}")
        if not os.path.exists(p):
            return p, letter
    raise RuntimeError(f"{td} は A-Z すべて使い切りました")

DATA_ROWS       = 10000
# グラフ 17cm（≈ 32行）が A6 から下に伸びる → ヘッダを十分下げて被り回避
HEADER_ROW_BASE = 44                    # 大見出し（4行ぶんの余裕）
START_ROW       = HEADER_ROW_BASE + 3   # 47
OUTLINE_ROWS    = 31
OUTLINE_START   = 8                     # S8〜T38

# ════════════════════════════════════════
# 計算方式：AISIN三角関数式 + 自動中心化
# ════════════════════════════════════════
# 目標：Sysmac の Act.Pos（0-360°回転エンコーダ値）を、
#   AISIN の4節リンク式で物理変位（mm）に変換し、
#   丸角矩形状の軌跡（X=±5、Y=±15 程度）にプロットする。
#
# 式：X = R·cos((θ−offset)) + √(L1²−(R·sin((θ−offset)))²) − L1
#   L1, R は AISIN の LX!C4/C7 (S/U軸用) と LZ!C4/C7 (LZ軸用) を参照
#   offset = データの平均（自動中心化）で θ を 0° 中心に正規化
CALC_METHOD = 'AISIN_TRIG_CENTERED'

SWING_ANGLE_DEG = 40  # 参考情報

# 色
COLOR_CMD     = '4472C4'  # 青
COLOR_ACT     = 'ED7D31'  # 橙
COLOR_OUTLINE = '548235'  # 濃緑

# ════════════════════════════════════════
# スタイル
# ════════════════════════════════════════
FONT_TITLE   = Font(name='游ゴシック', size=14, bold=True, color='FFFFFF')
FONT_SECTION = Font(name='游ゴシック', size=12, bold=True, color='FFFFFF')
FONT_HEADER  = Font(name='游ゴシック', size=10, bold=True, color='FFFFFF')
FONT_SUBHEAD = Font(name='游ゴシック', size=10, bold=True)
FONT_NORMAL  = Font(name='游ゴシック', size=10)
FONT_NOTE    = Font(name='游ゴシック', size=9, italic=True, color='666666')
FONT_SIGNAL  = Font(name='游ゴシック', size=9, color='1F4E78')

FILL_TITLE      = PatternFill('solid', fgColor='1F4E78')
FILL_SECT_LEFT  = PatternFill('solid', fgColor='2E75B6')
FILL_SECT_RIGHT = PatternFill('solid', fgColor='C55A11')
FILL_HEADER_L   = PatternFill('solid', fgColor='5B9BD5')
FILL_HEADER_R   = PatternFill('solid', fgColor='ED7D31')
FILL_CMD_IN     = PatternFill('solid', fgColor='DEEBF7')   # 薄青＝入力（指令）
FILL_ACT_IN     = PatternFill('solid', fgColor='FCE4D6')   # 薄橙＝入力（実測）
FILL_READONLY   = PatternFill('solid', fgColor='D9D9D9')   # 灰色＝入力不可（計算・自動・インデックス）
FILL_PARAM      = PatternFill('solid', fgColor='E2EFDA')
FILL_OUTLINE_BG = PatternFill('solid', fgColor='EAF3E4')

BORDER_THIN = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ════════════════════════════════════════
# ヘルパー
# ════════════════════════════════════════
def xform(src_cell, axis_sheet):
    """角度 [°] → 物理変位 [mm] 変換（AISIN LX/LZ シート同式）
       X = R·cosθ + √(L1² − (R·sinθ)²) − L1"""
    return (f'{axis_sheet}!$C$7*COS(RADIANS({src_cell}))'
            f'+SQRT({axis_sheet}!$C$4^2-({axis_sheet}!$C$7*SIN(RADIANS({src_cell})))^2)'
            f'-{axis_sheet}!$C$4')


def style_series(series, color_hex, line_width=19050, marker_symbol='circle', marker_size=3, dash=None):
    line = LineProperties(w=line_width, solidFill=color_hex, prstDash=dash if dash else 'solid')
    series.graphicalProperties = GraphicalProperties(ln=line)
    if marker_symbol == 'none' or marker_size == 0:
        # マーカー非表示（線のみ、10000点打つと重いため大きなデータセット用）
        m = Marker(symbol='none')
    else:
        m = Marker(symbol=marker_symbol, size=max(2, marker_size))
        m.graphicalProperties = GraphicalProperties(solidFill=color_hex)
        m.graphicalProperties.line = LineProperties(solidFill=color_hex)
    series.marker = m


def extract_existing_input(wb):
    """既存の入力データ（C,D,E,F,N,O 列）を dict で保存。
       データ行の開始位置はファイル内のヘッダ検索で自動判定。
       （旧版 START_ROW=39、新版 START_ROW=47 など変わり得るため）"""
    data = {}
    if SHEET_NEW not in wb.sheetnames:
        return data
    ws = wb[SHEET_NEW]

    # ファイル内のデータ開始行を自動検出：
    # A列に 1 が入っている最初の行（STEP=1）がデータ開始行
    detected_start = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v == 1:
            detected_start = r
            break
    if detected_start is None:
        print("  既存入力データ：STEP=1 の行が見つからないためスキップ")
        return data
    print(f"  既存ファイルのデータ開始行を検出: row {detected_start}")

    input_cols = [3, 4, 5, 6, 14, 15]   # C, D, E, F, N, O
    for col in input_cols:
        data[col] = {}
        for r in range(detected_start, detected_start + DATA_ROWS):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith('='):
                continue
            # rはファイル内行番号。STEP（A列）を基準にした相対番号（step_index）で保存
            step_index = r - detected_start  # 0〜DATA_ROWS-1
            data[col][step_index] = v
    total = sum(len(d) for d in data.values())
    print(f"  既存入力データ {total} セル分を抽出（{[(c, len(d)) for c, d in data.items()]}）")
    return data


def restore_input_data(ws, data):
    """抽出した入力値を新規シートへ復元"""
    count = 0
    for col, row_vals in data.items():
        for r, v in row_vals.items():
            ws.cell(row=r, column=col).value = v
            count += 1
    print(f"  入力データ {count} セル分を復元")


# ════════════════════════════════════════
# シート構築
# ════════════════════════════════════════
def create_slm8000_sheet(wb):
    if SHEET_NEW in wb.sheetnames:
        print(f"  既存シート '{SHEET_NEW}' を削除")
        del wb[SHEET_NEW]
    ws = wb.create_sheet(SHEET_NEW)
    ws.sheet_properties.tabColor = 'ED7D31'

    # ────────────────────────────────
    # 列幅（信号名 "MC_Axis_LZ.Cmd.Pos" = 18文字が隠れないよう広げる）
    # ────────────────────────────────
    col_widths = {
        # 左テーブル（S軸＋LZ軸）
        'A': 6,  'B': 9,
        'C': 20, 'D': 20, 'E': 20, 'F': 20,
        'G': 14, 'H': 14, 'I': 14, 'J': 14,
        # 仕切り
        'K': 3,
        # 右テーブル（U軸のみ、LZは削除）
        'L': 6,  'M': 9,
        'N': 20, 'O': 20,
        'P': 14, 'Q': 14,
        # 仕切り
        'R': 3,
        # 製品輪郭
        'S': 12, 'T': 12,
        # 仕切り
        'U': 3,
        # パラメータ
        'V': 18, 'W': 13, 'X': 30,
    }
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    # ────────────────────────────────
    # タイトル・運用説明（行1〜4）
    # ────────────────────────────────
    ws.merge_cells('A1:X1')
    ws['A1'].value = 'SLM8000方式 軌道計算シート（Omron Sysmac Studio データトレース用・左右軸対応）'
    ws['A1'].font = FONT_TITLE; ws['A1'].fill = FILL_TITLE; ws['A1'].alignment = ALIGN_C
    ws.row_dimensions[1].height = 28

    notes = [
        '【運用手順】① Sysmac Studio → データトレース実行　② CSV 出力　'
        '③ 下段テーブルへコピペ：左テーブル C/D=S軸、E/F=LZ軸、右テーブル N/O=U軸　'
        '（LZ軸は左右共通のため左テーブルへのみ入力）',
        '【計算方式：AISIN三角関数式＋自動中心化】'
        'X = R·cos(θ−offset) + √(L1²−(R·sin(θ−offset))²) − L1　（L1, Rは AISIN LX/LZ シートから継承）。'
        'offset はデータ平均で自動中心化 → トレース値を0°中心に正規化して丸角矩形の軌跡を描画。',
        '【グラフ凡例】青=軌跡（実測 Act.Pos から計算）　赤=製品輪郭（AISIN既存ワーク寸法）。'
        'グリッド：主5mm／副1mm　軸範囲：±20mm',
    ]
    for i, t in enumerate(notes):
        r = i + 2
        ws.merge_cells(f'A{r}:X{r}')
        ws[f'A{r}'].value = t
        ws[f'A{r}'].font = FONT_NOTE
        ws[f'A{r}'].alignment = ALIGN_L
        ws.row_dimensions[r].height = 22

    # ────────────────────────────────
    # パラメータ部（V〜X、行6〜14）
    # ────────────────────────────────
    ws.merge_cells('V6:X6')
    ws['V6'].value = 'パラメータ（AISIN参照）'
    ws['V6'].font = FONT_SECTION; ws['V6'].fill = FILL_TITLE; ws['V6'].alignment = ALIGN_C
    ws.row_dimensions[6].height = 22

    # ════════════════════════════════════════
    # パラメータ表（AISIN三角関数式 + 自動中心化）
    # AISIN LX!C4/C7, LZ!C4/C7 を参照、オフセット自動 = AVERAGE
    # ════════════════════════════════════════
    _end_row = START_ROW + DATA_ROWS - 1
    params = [
        ('計算方式',            'AISIN三角関数＋自動中心化',       'X = R·cos(θ-off) + √(L1²-(R·sinθ)²) - L1'),
        ('LX L1 [mm]',          '=LX!C4',                          'S/U軸 基準腕長 (AISIN参照)'),
        ('LX R [mm]',           '=LX!C7',                          'S/U軸 振幅半径 (AISIN参照)'),
        ('LZ L1 [mm]',          '=LZ!C4',                          'LZ軸 基準腕長 (AISIN参照)'),
        ('LZ R [mm]',           '=LZ!C7',                          'LZ軸 振幅半径 (AISIN参照)'),
        ('S 軸 オフセット [°]', f'=IFERROR(AVERAGE(C{START_ROW}:D{_end_row}),0)',
                                 'S.Cmd/Act の平均で自動中心化（手動上書き可）'),
        ('U 軸 オフセット [°]', f'=IFERROR(AVERAGE(N{START_ROW}:O{_end_row}),0)',
                                 'U.Cmd/Act の平均で自動中心化'),
        ('LZ軸 オフセット [°]', f'=IFERROR(AVERAGE(E{START_ROW}:F{_end_row}),0)',
                                 'LZ.Cmd/Act の平均で自動中心化'),
        ('振り角 [°]',          SWING_ANGLE_DEG,                   '（操作指定・参考）'),
        ('巻方向',              '=LZ!J16',                         'R巻／L巻'),
        ('マスタ軸',            'MC_Axis_Winding',                 '0-360°/cycle (参考)'),
        ('カム表',              'Wind_SU_R',                       '0-40mm 台形 (参考)'),
    ]
    for i, (label, value, note) in enumerate(params):
        r = 7 + i
        ws.cell(row=r, column=22).value = label
        ws.cell(row=r, column=22).font = FONT_SUBHEAD
        ws.cell(row=r, column=22).fill = FILL_PARAM
        ws.cell(row=r, column=22).border = BORDER_THIN
        ws.cell(row=r, column=22).alignment = ALIGN_L
        ws.cell(row=r, column=23).value = value
        ws.cell(row=r, column=23).font = FONT_NORMAL
        ws.cell(row=r, column=23).fill = FILL_PARAM
        ws.cell(row=r, column=23).border = BORDER_THIN
        ws.cell(row=r, column=23).alignment = ALIGN_C
        ws.cell(row=r, column=24).value = note
        ws.cell(row=r, column=24).font = FONT_NOTE
        ws.cell(row=r, column=24).fill = FILL_PARAM
        ws.cell(row=r, column=24).border = BORDER_THIN
        ws.cell(row=r, column=24).alignment = ALIGN_L
        ws.row_dimensions[r].height = 18

    # ────────────────────────────────
    # 製品輪郭データ（S〜T、行6〜38）── 等加速度20230825!F6:G36 を参照
    # ────────────────────────────────
    ws.merge_cells('S6:T6')
    ws['S6'].value = '製品輪郭（参照）'
    ws['S6'].font = FONT_SECTION; ws['S6'].fill = FILL_TITLE; ws['S6'].alignment = ALIGN_C

    ws['S7'].value = 'X [mm]'; ws['T7'].value = 'Y [mm]'
    for addr in ('S7', 'T7'):
        ws[addr].font = FONT_SUBHEAD; ws[addr].fill = FILL_OUTLINE_BG
        ws[addr].alignment = ALIGN_C; ws[addr].border = BORDER_THIN

    for i in range(OUTLINE_ROWS):
        src_row = 6 + i
        dst_row = OUTLINE_START + i
        ws.cell(row=dst_row, column=19).value = f"='等加速度20230825'!F{src_row}"
        ws.cell(row=dst_row, column=20).value = f"='等加速度20230825'!G{src_row}"
        for col in (19, 20):
            c = ws.cell(row=dst_row, column=col)
            c.border = BORDER_THIN
            c.fill = FILL_OUTLINE_BG
            c.alignment = ALIGN_C

    # ────────────────────────────────
    # 空白スペーサ（行 36〜43）
    # チャート（17cm ≈ 32行）が A6 から下に伸びるため、
    # テーブルヘッダをこれ以上下げることで被りを回避
    # ────────────────────────────────
    for r_spacer in range(36, HEADER_ROW_BASE):
        ws.row_dimensions[r_spacer].height = 12  # 詰めて表示

    # ────────────────────────────────
    # データテーブル 大見出し（行 HEADER_ROW_BASE = 44）
    # ────────────────────────────────
    r_big = HEADER_ROW_BASE  # 44
    ws.merge_cells(start_row=r_big, start_column=1, end_row=r_big, end_column=10)
    ws.cell(row=r_big, column=1).value = '◆ 左側：S軸（振り）＋ LZ軸（上下） データ入力 & 計算'
    ws.cell(row=r_big, column=1).font = FONT_SECTION
    ws.cell(row=r_big, column=1).fill = FILL_SECT_LEFT
    ws.cell(row=r_big, column=1).alignment = ALIGN_C

    ws.merge_cells(start_row=r_big, start_column=12, end_row=r_big, end_column=17)
    ws.cell(row=r_big, column=12).value = '◆ 右側：U軸（振り） データ入力 & 計算  ※LZ軸は左テーブル参照'
    ws.cell(row=r_big, column=12).font = FONT_SECTION
    ws.cell(row=r_big, column=12).fill = FILL_SECT_RIGHT
    ws.cell(row=r_big, column=12).alignment = ALIGN_C
    ws.row_dimensions[r_big].height = 24

    # ────────────────────────────────
    # 中見出し（行37）
    # ────────────────────────────────
    r_mid = r_big + 1   # 37
    mid_headers = [
        # 左テーブル
        (1, 1, 'STEP',                 FILL_HEADER_L),
        (2, 2, '時刻[ms]',             FILL_HEADER_L),
        (3, 4, 'S軸（振り）入力 [°]',  FILL_HEADER_L),
        (5, 6, 'LZ軸（上下）入力 [mm]', FILL_HEADER_L),
        (7, 8, 'S軸 → X変位 [mm]',     FILL_HEADER_L),
        (9,10, 'LZ軸 → Z変位 [mm]',    FILL_HEADER_L),
        # 右テーブル（U軸のみ、LZは撤去）
        (12,12,'STEP',                 FILL_HEADER_R),
        (13,13,'時刻[ms]',             FILL_HEADER_R),
        (14,15,'U軸（振り）入力 [°]',  FILL_HEADER_R),
        (16,17,'U軸 → X変位 [mm]',     FILL_HEADER_R),
    ]
    for c_start, c_end, label, fill in mid_headers:
        if c_start == c_end:
            cell = ws.cell(row=r_mid, column=c_start)
        else:
            ws.merge_cells(start_row=r_mid, start_column=c_start, end_row=r_mid, end_column=c_end)
            cell = ws.cell(row=r_mid, column=c_start)
        cell.value = label
        cell.font = FONT_HEADER; cell.fill = fill
        cell.alignment = ALIGN_C; cell.border = BORDER_THIN
    ws.row_dimensions[r_mid].height = 28

    # ────────────────────────────────
    # 信号名ヘッダ（行38）
    # ────────────────────────────────
    r_sig = r_mid + 1   # 38
    sig_headers = [
        # 左
        (1,  '#',                    FILL_READONLY),
        (2,  '(連番)',               FILL_READONLY),
        (3,  'MC_Axis_S.Cmd.Pos',    FILL_CMD_IN),
        (4,  'MC_Axis_S.Act.Pos',    FILL_ACT_IN),
        (5,  'MC_Axis_LZ.Cmd.Pos',   FILL_CMD_IN),
        (6,  'MC_Axis_LZ.Act.Pos',   FILL_ACT_IN),
        (7,  'S.Cmd → X変位',        FILL_READONLY),
        (8,  'S.Act → X変位',        FILL_READONLY),
        (9,  'LZ.Cmd → Z変位',       FILL_READONLY),
        (10, 'LZ.Act → Z変位',       FILL_READONLY),
        # 右（LZ重複列なし）
        (12, '#',                    FILL_READONLY),
        (13, '(連番)',               FILL_READONLY),
        (14, 'MC_Axis_U.Cmd.Pos',    FILL_CMD_IN),
        (15, 'MC_Axis_U.Act.Pos',    FILL_ACT_IN),
        (16, 'U.Cmd → X変位',        FILL_READONLY),
        (17, 'U.Act → X変位',        FILL_READONLY),
    ]
    for col, label, fill in sig_headers:
        c = ws.cell(row=r_sig, column=col)
        c.value = label; c.font = FONT_SIGNAL
        c.fill = fill; c.alignment = ALIGN_C; c.border = BORDER_THIN
    ws.row_dimensions[r_sig].height = 22

    # コメント（入力列への手引き）
    ws.cell(row=r_sig, column=3).comment = Comment(
        'Sysmac Studio データトレースで取得した MC_Axis_S.Cmd.Pos（S軸指令 [°]）をここから下にコピペ', 'SLM8000')
    ws.cell(row=r_sig, column=5).comment = Comment(
        'Sysmac Studio データトレースで取得した MC_Axis_LZ.Cmd.Pos（LZ軸指令 [mm]）をここから下にコピペ', 'SLM8000')
    ws.cell(row=r_sig, column=14).comment = Comment(
        'Sysmac Studio データトレースで取得した MC_Axis_U.Cmd.Pos（U軸指令 [°]）をここから下にコピペ', 'SLM8000')

    # ────────────────────────────────
    # データ行（行39〜10038）
    # ────────────────────────────────
    # ────────────────────────────────
    # AISIN三角関数式（自動中心化）：
    #   X = R·cos(θ-offset) + √(L1² - (R·sin(θ-offset))²) - L1
    #   θ-offset で 0°中心に正規化してから4節リンク式で mm 変位化
    # パラメータ参照：
    #   $W$8=LX!C4=162, $W$9=LX!C7=5.1, $W$10=LZ!C4=130, $W$11=LZ!C7=15
    #   $W$12=S軸オフセット（自動）, $W$13=U軸オフセット, $W$14=LZ軸オフセット
    # ────────────────────────────────
    def fS(src, off):   # S/U軸 LX系
        return (f'$W$9*COS(RADIANS({src}-{off}))'
                f'+SQRT($W$8^2-($W$9*SIN(RADIANS({src}-{off})))^2)'
                f'-$W$8')
    def fLZ(src, off):  # LZ軸 LZ系
        return (f'$W$11*COS(RADIANS({src}-{off}))'
                f'+SQRT($W$10^2-($W$11*SIN(RADIANS({src}-{off})))^2)'
                f'-$W$10')

    print(f"  データ行 {DATA_ROWS} 行ぶん（AISIN三角関数＋自動中心化）の数式を書き込み中"
          f"（row {START_ROW} 〜 {START_ROW + DATA_ROWS - 1}）...")
    for i in range(DATA_ROWS):
        r = START_ROW + i
        # 左テーブル
        ws.cell(row=r, column=1).value = i + 1;  ws.cell(row=r, column=1).fill = FILL_READONLY
        ws.cell(row=r, column=2).value = i;      ws.cell(row=r, column=2).fill = FILL_READONLY
        ws.cell(row=r, column=3).fill = FILL_CMD_IN
        ws.cell(row=r, column=4).fill = FILL_ACT_IN
        ws.cell(row=r, column=5).fill = FILL_CMD_IN
        ws.cell(row=r, column=6).fill = FILL_ACT_IN
        # S軸計算（LX系）、オフセット=$W$12
        ws.cell(row=r, column=7).value  = f'=IF(C{r}="","",{fS(f"C{r}","$W$12")})'
        ws.cell(row=r, column=8).value  = f'=IF(D{r}="","",{fS(f"D{r}","$W$12")})'
        # LZ軸計算（LZ系）、オフセット=$W$14
        ws.cell(row=r, column=9).value  = f'=IF(E{r}="","",{fLZ(f"E{r}","$W$14")})'
        ws.cell(row=r, column=10).value = f'=IF(F{r}="","",{fLZ(f"F{r}","$W$14")})'
        for col in range(7, 11):
            ws.cell(row=r, column=col).fill = FILL_READONLY

        # 右テーブル
        ws.cell(row=r, column=12).value = f'=A{r}'; ws.cell(row=r, column=12).fill = FILL_READONLY
        ws.cell(row=r, column=13).value = f'=B{r}'; ws.cell(row=r, column=13).fill = FILL_READONLY
        ws.cell(row=r, column=14).fill = FILL_CMD_IN
        ws.cell(row=r, column=15).fill = FILL_ACT_IN
        # U軸計算（LX系）、オフセット=$W$13
        ws.cell(row=r, column=16).value = f'=IF(N{r}="","",{fS(f"N{r}","$W$13")})'
        ws.cell(row=r, column=17).value = f'=IF(O{r}="","",{fS(f"O{r}","$W$13")})'
        for col in (16, 17):
            ws.cell(row=r, column=col).fill = FILL_READONLY

        if i < 200 or i % 1000 == 0:
            for col in list(range(1, 11)) + list(range(12, 18)):
                ws.cell(row=r, column=col).border = BORDER_THIN

    # ────────────────────────────────
    # グラフ共通ビルダ
    # ────────────────────────────────
    # ────────────────────────────────
    # XY軌跡チャート：青=軌跡、赤=製品輪郭、±20mm固定範囲・細かいグリッド
    # ユーザー指定のターゲット画像（軌跡 / 製品 / 細かいグリッド）を再現
    # ────────────────────────────────
    def build_trajectory_chart(title, x_cmd_ref, y_cmd_ref, x_act_ref, y_act_ref,
                               x_out_ref, y_out_ref):
        chart = ScatterChart()
        chart.title = title
        chart.scatterStyle = 'lineMarker'
        chart.style = 13
        chart.x_axis.title = 'X方向 [mm]'
        chart.y_axis.title = 'Z方向 [mm]'
        # 範囲 ±20mm 固定（ターゲット画像に合わせる）
        chart.x_axis.scaling.min = -20; chart.x_axis.scaling.max = 20
        chart.y_axis.scaling.min = -20; chart.y_axis.scaling.max = 20
        # 目盛：主5mm・副1mm、整数表示、グリッド両方表示
        chart.x_axis.majorUnit = 5; chart.x_axis.minorUnit = 1
        chart.y_axis.majorUnit = 5; chart.y_axis.minorUnit = 1
        chart.x_axis.majorTickMark = 'out'; chart.x_axis.minorTickMark = 'out'
        chart.y_axis.majorTickMark = 'out'; chart.y_axis.minorTickMark = 'out'
        chart.x_axis.number_format = '0'; chart.y_axis.number_format = '0'
        chart.x_axis.majorGridlines = ChartLines()
        chart.y_axis.majorGridlines = ChartLines()
        chart.x_axis.minorGridlines = ChartLines()
        chart.y_axis.minorGridlines = ChartLines()

        chart.width  = 22
        chart.height = 17
        chart.legend.position = 'r'
        chart.legend.overlay = False

        # 青：実測軌跡（Act → トレース本体、メインで表示）
        s_act = Series(y_act_ref, x_act_ref, title='軌跡（実測）')
        style_series(s_act, '4472C4', line_width=25400, marker_symbol='circle', marker_size=0)
        chart.series.append(s_act)

        # 青（薄）：指令軌跡（Cmd → 参考）
        s_cmd = Series(y_cmd_ref, x_cmd_ref, title='軌跡（指令）')
        style_series(s_cmd, '8EA9DB', line_width=12700, marker_symbol='circle', marker_size=0, dash='sysDash')
        chart.series.append(s_cmd)

        # 赤：製品輪郭（AISIN既存ワーク）
        s_out = Series(y_out_ref, x_out_ref, title='製品')
        style_series(s_out, 'C00000', line_width=31750, marker_symbol='circle', marker_size=0)
        chart.series.append(s_out)

        return chart

    # 製品輪郭参照（S/T列）
    x_out_ref = Reference(ws, min_col=19, max_col=19,
                          min_row=OUTLINE_START, max_row=OUTLINE_START + OUTLINE_ROWS - 1)
    y_out_ref = Reference(ws, min_col=20, max_col=20,
                          min_row=OUTLINE_START, max_row=OUTLINE_START + OUTLINE_ROWS - 1)

    # 左グラフ：X=G(S.Cmd→X)/H(S.Act→X)、Y=I(LZ.Cmd→Z)/J(LZ.Act→Z)
    x_L_cmd = Reference(ws, min_col=7,  max_col=7,  min_row=START_ROW, max_row=START_ROW + DATA_ROWS - 1)
    y_L_cmd = Reference(ws, min_col=9,  max_col=9,  min_row=START_ROW, max_row=START_ROW + DATA_ROWS - 1)
    x_L_act = Reference(ws, min_col=8,  max_col=8,  min_row=START_ROW, max_row=START_ROW + DATA_ROWS - 1)
    y_L_act = Reference(ws, min_col=10, max_col=10, min_row=START_ROW, max_row=START_ROW + DATA_ROWS - 1)
    chart_left = build_trajectory_chart(
        '【左側】S軸 巻線軌跡 ＆ 製品輪郭',
        x_L_cmd, y_L_cmd, x_L_act, y_L_act, x_out_ref, y_out_ref,
    )
    ws.add_chart(chart_left, 'A6')

    # 右グラフ：X=P(U.Cmd→X)/Q(U.Act→X)、Y=I/J（LZは左から参照）
    x_R_cmd = Reference(ws, min_col=16, max_col=16, min_row=START_ROW, max_row=START_ROW + DATA_ROWS - 1)
    y_R_cmd = Reference(ws, min_col=9,  max_col=9,  min_row=START_ROW, max_row=START_ROW + DATA_ROWS - 1)
    x_R_act = Reference(ws, min_col=17, max_col=17, min_row=START_ROW, max_row=START_ROW + DATA_ROWS - 1)
    y_R_act = Reference(ws, min_col=10, max_col=10, min_row=START_ROW, max_row=START_ROW + DATA_ROWS - 1)
    chart_right = build_trajectory_chart(
        '【右側】U軸 巻線軌跡 ＆ 製品輪郭',
        x_R_cmd, y_R_cmd, x_R_act, y_R_act, x_out_ref, y_out_ref,
    )
    ws.add_chart(chart_right, 'N6')

    # フリーズペイン（データ入力列から固定）
    ws.freeze_panes = ws.cell(row=START_ROW, column=3)

    # メタ
    footer_row = START_ROW + DATA_ROWS + 2
    ws.cell(row=footer_row, column=1).value = (
        f'生成日時: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}  '
        f'生成元: append_slm8000_method.py (v3)  '
        f'軸: 左=S軸／右=U軸／LZ軸は左右共通'
    )
    ws.cell(row=footer_row, column=1).font = FONT_NOTE

    return ws


# ════════════════════════════════════════
# エントリポイント
# ════════════════════════════════════════
def main():
    # 入力：最新の作業ファイル
    input_path = find_latest_input()
    if not os.path.exists(input_path):
        print(f"ERROR: 入力ファイルが見つかりません: {input_path}", file=sys.stderr)
        sys.exit(1)
    print(f"入力: {input_path}")

    # 出力：今日の次の空きラベル（命名規則 {BASE}_{YYYYMMDD}{A-Z}.xlsx）
    output_path, letter = next_output_path()
    print(f"出力: {output_path}  （今日{letter}版）")

    wb = load_workbook(input_path)
    print(f"既存シート: {wb.sheetnames}")

    # 既存入力データを抽出
    saved_input = extract_existing_input(wb)

    # 非回帰チェック用
    before_charts = {sn: len(wb[sn]._charts) for sn in wb.sheetnames if sn != SHEET_NEW}

    # シート生成
    create_slm8000_sheet(wb)

    # 入力データを復元：抽出時は step_index（0-origin）で保存したので、
    # 新 START_ROW を足して新ファイル上の行番号に変換する
    new_saved = {}
    for col, step_vals in saved_input.items():
        new_saved[col] = {}
        for step_index, v in step_vals.items():
            new_r = START_ROW + step_index
            if START_ROW <= new_r < START_ROW + DATA_ROWS:
                new_saved[col][new_r] = v
    ws_new = wb[SHEET_NEW]
    restore_input_data(ws_new, new_saved)

    # 保存
    try:
        wb.save(output_path)
        print(f"保存完了: {output_path}")
    except PermissionError:
        print(f"\nERROR: 出力先が開かれています: {output_path}", file=sys.stderr)
        sys.exit(1)

    # 非回帰チェック
    print("\n--- 非回帰チェック ---")
    wb2 = load_workbook(output_path)
    for sn, before_n in before_charts.items():
        after_n = len(wb2[sn]._charts)
        status = 'OK' if before_n == after_n else 'NG'
        print(f"  [{status}] {sn}: charts {before_n} → {after_n}")
    new_charts = len(wb2[SHEET_NEW]._charts)
    print(f"  新規シート charts = {new_charts}  (期待値: 2)")
    for i, ch in enumerate(wb2[SHEET_NEW]._charts):
        ext_cx = ch.anchor.ext.cx / 360000 if hasattr(ch.anchor, 'ext') and ch.anchor.ext else 'n/a'
        ext_cy = ch.anchor.ext.cy / 360000 if hasattr(ch.anchor, 'ext') and ch.anchor.ext else 'n/a'
        fr = getattr(ch.anchor, '_from', None) or getattr(ch.anchor, 'from_', None)
        print(f"    chart[{i}] series={len(ch.series)}  "
              f"anchor=col{fr.col},row{fr.row}  "
              f"ext={ext_cx:.1f}×{ext_cy:.1f}cm  "
              f"legendPos={ch.legend.position}")

    # 入力データ保持確認（新行位置）
    print(f"\n--- 入力データ保持チェック（row {START_ROW}, {START_ROW+961}, {START_ROW+DATA_ROWS-1}） ---")
    ws_chk = wb2[SHEET_NEW]
    for r in (START_ROW, START_ROW+961, START_ROW+DATA_ROWS-1):
        vals = [ws_chk.cell(row=r, column=c).value for c in (3, 4, 5, 6, 14, 15)]
        print(f"  row{r}: C={vals[0]}  D={vals[1]}  E={vals[2]}  F={vals[3]}  N={vals[4]}  O={vals[5]}")

    # チャートの下端行を計算（被りの目視確認用）
    # 17cm = 482pt。 行1-5合計=109pt。Row 6-13 合計=22+18*7=148pt。以降default 15pt。
    # → chart 下端は row 6+約32 ≈ row 38付近。HEADER_ROW_BASE=44 なので 6行バッファ。
    print(f"\n  チャート下端推定: row 38 付近  / テーブル開始: row {HEADER_ROW_BASE}  → バッファ約 {HEADER_ROW_BASE-38} 行")
    print("\n完了")


if __name__ == '__main__':
    main()

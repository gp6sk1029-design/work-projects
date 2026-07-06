# -*- coding: utf-8 -*-
"""
【パッチ用スクリプト】データ入力エリアを明示化

【目的】
  NAKAM シートの col 1-4（θ指令/機械、Z指令/機械）に既存の SLM8000 サンプル
  データが入っており、ユーザーが自分のデータを貼り付ける位置が分かりにくい。
  既存データをクリアし、色分けと注記で入力エリアを明確化する。

【変更内容】
  1. col 1-4 の既存データをクリア（row 4 以降）※数式 col 5-12 は保持
  2. 色分け：
     - col 1 (θ指令) / col 3 (Z指令) = 薄青（Cmd系）
     - col 2 (θ機械) / col 4 (Z機械) = 薄橙（Act系、×100 適用対象）
  3. row 3 の信号名ヘッダに Sysmac Studio 信号名を明記
  4. コメント追加：「Sysmac Studio データトレース値をここに貼り付け」
  5. 数式範囲（col 5-12）を全データ行分保持

【入力】 命名規則で最新の AISIN反映版
【出力】 次のレター版
"""

import sys
import os
import string
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

DESKTOP = r"C:\Users\SEIGI-N13\Desktop"
BASE_NAME = "【巻線検証】SLM8000軌道_AISIN反映"


def find_latest_input():
    td = datetime.now().strftime('%Y%m%d')
    for letter in reversed(string.ascii_uppercase):
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if os.path.exists(p):
            return p
    raise RuntimeError("AISIN反映ファイルが見つかりません")


def next_output_path():
    td = datetime.now().strftime('%Y%m%d')
    for letter in string.ascii_uppercase:
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if not os.path.exists(p):
            return p, letter
    raise RuntimeError(f"{td} は A-Z すべて使い切りました")


# スタイル
FILL_CMD    = PatternFill('solid', fgColor='DEEBF7')   # 薄青：指令（Cmd）
FILL_ACT    = PatternFill('solid', fgColor='FCE4D6')   # 薄橙：実測（Act、×100対象）
FILL_HEADER = PatternFill('solid', fgColor='1F4E78')   # 濃紺
FONT_HEADER = Font(name='游ゴシック', size=10, bold=True, color='FFFFFF')
FONT_SIGNAL = Font(name='游ゴシック', size=9, color='1F4E78', italic=True)
FONT_TITLE  = Font(name='游ゴシック', size=11, bold=True, color='FFFFFF')
ALIGN_C     = Alignment(horizontal='center', vertical='center', wrap_text=True)
BORDER_THIN = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)


def apply_patch(wb):
    ws = wb['NAKAM']
    print(f"  NAKAM max_row = {ws.max_row}")

    # ─────────────────────────
    # 1. 既存のサンプルデータをクリア（row 4 以降、col 1-4）
    # ─────────────────────────
    cleared = 0
    for r in range(4, ws.max_row + 1):
        for c in range(1, 5):   # col 1-4
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None
                cleared += 1
    print(f"  サンプルデータをクリア: {cleared} セル")

    # ─────────────────────────
    # 2. ヘッダ更新（row 1-3）：信号名を Sysmac Studio 形式に
    # ─────────────────────────
    # row 1 は既存の 'データ代入' ヘッダ（セル結合されている可能性）
    # row 2 は 'θ' / 'Z'
    # row 3 は '指令位置' / '機械位置'
    # → row 3 に Sysmac 信号名を追記（改行で）

    sig_names = {
        1: 'θ指令位置\nMC_Axis_θ.Cmd.Pos',
        2: 'θ機械位置 (×100)\nMC_Axis_θ.Act.Pos',
        3: 'Z指令位置\nMC_Axis_Z.Cmd.Pos',
        4: 'Z機械位置 (×100)\nMC_Axis_Z.Act.Pos',
    }
    # row 3 に書き込み
    for col, name in sig_names.items():
        cell = ws.cell(row=3, column=col)
        cell.value = name
        cell.font = FONT_SIGNAL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN
    ws.row_dimensions[3].height = 36
    print(f"  row 3 ヘッダを Sysmac 信号名に更新")

    # ─────────────────────────
    # 3. 入力エリア（row 4 以降 col 1-4）に色分け・罫線を適用
    # ─────────────────────────
    # どこまで色付けするか？ → F 列（グラフ計算式）が存在する行まで
    # F 列の数式が存在する最後の行を特定
    last_formula_row = 3
    for r in range(4, ws.max_row + 1):
        f_val = ws.cell(row=r, column=6).value
        if f_val is not None and isinstance(f_val, str) and f_val.startswith('='):
            last_formula_row = r
    print(f"  F列計算式の最終行: {last_formula_row}（この範囲まで入力セル色付け）")

    for r in range(4, last_formula_row + 1):
        # col 1, 3 = Cmd（薄青）、col 2, 4 = Act（薄橙）
        for col, fill in ((1, FILL_CMD), (2, FILL_ACT), (3, FILL_CMD), (4, FILL_ACT)):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.border = BORDER_THIN
    print(f"  入力セル色付け: col 1-4, row 4〜{last_formula_row}")

    # ─────────────────────────
    # 4. コメント：入力エリアへの注記
    # ─────────────────────────
    ws.cell(row=4, column=1).comment = Comment(
        'Sysmac Studio データトレースで取得した MC_Axis_θ.Cmd.Pos（θ指令位置）を、\n'
        'ここから下に貼り付けてください。\n'
        '※ ×100 補正なし（そのまま）',
        'AISIN反映版'
    )
    ws.cell(row=4, column=2).comment = Comment(
        'Sysmac Studio データトレースで取得した MC_Axis_θ.Act.Pos（θ機械位置）を、\n'
        'ここから下に貼り付けてください。\n'
        '※ ×100 補正あり（グラフ計算で自動適用）',
        'AISIN反映版'
    )
    ws.cell(row=4, column=3).comment = Comment(
        'Sysmac Studio データトレースで取得した MC_Axis_Z.Cmd.Pos（Z指令位置）を、\n'
        'ここから下に貼り付けてください。\n'
        '※ ×100 補正なし',
        'AISIN反映版'
    )
    ws.cell(row=4, column=4).comment = Comment(
        'Sysmac Studio データトレースで取得した MC_Axis_Z.Act.Pos（Z機械位置）を、\n'
        'ここから下に貼り付けてください。\n'
        '※ ×100 補正あり',
        'AISIN反映版'
    )
    print(f"  入力セルにコメント追加（col 1-4 row 4）")

    # ─────────────────────────
    # 5. row 1 の「データ代入」タイトル行を強調
    # ─────────────────────────
    # row 1 col 1-4 に色を付ける
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_TITLE
        cell.alignment = ALIGN_C
    # 値を上書き（結合セルでない場合）
    try:
        ws.cell(row=1, column=1).value = '◆ データ入力エリア（Sysmac Studio トレース値を貼り付け）'
    except Exception as e:
        print(f"  (row 1 タイトル更新スキップ: {e})")
    ws.row_dimensions[1].height = 26

    return cleared


def main():
    input_path = find_latest_input()
    output_path, letter = next_output_path()
    print(f"入力: {input_path}")
    print(f"出力: {output_path}  （今日{letter}版）")

    wb = load_workbook(input_path)
    print(f"\n--- パッチ適用中 ---")
    cleared = apply_patch(wb)

    print(f"\n--- 保存 ---")
    try:
        wb.save(output_path)
        print(f"完了: {output_path}")
    except PermissionError:
        print(f"ERROR: 出力先が開かれています: {output_path}", file=sys.stderr)
        sys.exit(1)

    # 検証
    print(f"\n--- 検証 ---")
    wb2 = load_workbook(output_path)
    ws = wb2['NAKAM']
    print(f"  NAKAM charts: {len(ws._charts)}")
    print(f"  row 1-3 col 1-4 サンプル:")
    for r in range(1, 4):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 5)]
        print(f"    row{r}: {row_vals}")
    print(f"  row 4-6 col 1-4 サンプル（空のはず）:")
    for r in range(4, 7):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 5)]
        print(f"    row{r}: {row_vals}")
    # 数式が壊れていないか確認
    print(f"  row 4 col 5-8（数式）:")
    for c in range(5, 9):
        v = ws.cell(row=4, column=c).value
        print(f"    col{c}: {v}")


if __name__ == '__main__':
    main()

# -*- coding: utf-8 -*-
"""
【パッチ用スクリプト】グラフの目盛を細かくして数値表示
- X/Y軸の主目盛=5mm刻み（数値表示あり）
- X/Y軸の副目盛=1mm刻み（目盛線のみ）
- 目盛線を外側に表示、数値フォーマットを '0' (整数)

データ・数式・他の設定は一切触らない。
実行前に Excel でファイルを閉じておくこと。
"""
import sys, os
from openpyxl import load_workbook

PATH = r"C:\Users\SEIGI-N13\Desktop\2976 LXLZ軌跡計算式_20230525_NEW.xlsx"
SHEET = "SLM8000方式_軌道計算"

# 目盛設定
MAJOR_UNIT = 5     # 主目盛（数値表示）の間隔 [mm]
MINOR_UNIT = 1     # 副目盛（線のみ）の間隔 [mm]
NUM_FORMAT = '0'   # 数値フォーマット（整数表示）
AXIS_MIN   = -20
AXIS_MAX   = 20

if not os.path.exists(PATH):
    print(f"ERROR: {PATH} が見つかりません", file=sys.stderr); sys.exit(1)

print(f"読み込み: {PATH}")
wb = load_workbook(PATH)
ws = wb[SHEET]
print(f"シート内チャート数: {len(ws._charts)}")

for i, ch in enumerate(ws._charts):
    # 軸スケール（念のため再設定）
    ch.x_axis.scaling.min = AXIS_MIN
    ch.x_axis.scaling.max = AXIS_MAX
    ch.y_axis.scaling.min = AXIS_MIN
    ch.y_axis.scaling.max = AXIS_MAX

    # 主目盛・副目盛 間隔
    ch.x_axis.majorUnit = MAJOR_UNIT
    ch.x_axis.minorUnit = MINOR_UNIT
    ch.y_axis.majorUnit = MAJOR_UNIT
    ch.y_axis.minorUnit = MINOR_UNIT

    # 目盛線を外側に表示
    ch.x_axis.majorTickMark = 'out'
    ch.x_axis.minorTickMark = 'out'
    ch.y_axis.majorTickMark = 'out'
    ch.y_axis.minorTickMark = 'out'

    # 数値フォーマット（整数）
    ch.x_axis.number_format = NUM_FORMAT
    ch.y_axis.number_format = NUM_FORMAT

    # 軸ラベルの表示位置を明示
    ch.x_axis.tickLblPos = 'nextTo'
    ch.y_axis.tickLblPos = 'nextTo'

    # グリッドを表示（副目盛にも）
    from openpyxl.chart.axis import ChartLines
    ch.x_axis.majorGridlines = ChartLines()
    ch.y_axis.majorGridlines = ChartLines()
    ch.x_axis.minorGridlines = ChartLines()
    ch.y_axis.minorGridlines = ChartLines()

    # タイトルは既存のまま保持（title属性へ手を加えない）
    print(f"  chart[{i}]: 軸設定更新 "
          f"scaling=[{AXIS_MIN},{AXIS_MAX}] major={MAJOR_UNIT} minor={MINOR_UNIT} fmt={NUM_FORMAT}")

try:
    wb.save(PATH)
    print(f"\n保存完了: {PATH}")
except PermissionError:
    print("\nERROR: Excel で開かれています。閉じて再実行してください。", file=sys.stderr)
    sys.exit(1)

# 非回帰確認
wb2 = load_workbook(PATH)
ws2 = wb2[SHEET]
print(f"\n--- 確認 ---")
print(f"シート数: {len(wb2.sheetnames)}  新規シート チャート数: {len(ws2._charts)}")
for i, ch in enumerate(ws2._charts):
    print(f"  chart[{i}] majorUnit={ch.x_axis.majorUnit}/{ch.y_axis.majorUnit}  "
          f"minorUnit={ch.x_axis.minorUnit}/{ch.y_axis.minorUnit}  "
          f"tick(X)={ch.x_axis.majorTickMark}/{ch.x_axis.minorTickMark}  "
          f"fmt={ch.x_axis.number_format}")

# 既存入力データの非回帰（壊れていないか）
print(f"\n--- 入力データの非回帰（row 10038 サンプル） ---")
for col, label in [(3, 'C(S.Cmd)'), (6, 'F(LZ.Act)'), (15, 'O(U.Act)')]:
    v = ws2.cell(row=10038, column=col).value
    print(f"  {label}: {v}")

# -*- coding: utf-8 -*-
"""
【パッチ用スクリプト】機械位置データに ×100 を適用

【目的】
  機械位置（B列=θ機械位置、D列=Z機械位置）に貼り付けられた Sysmac Studio
  のトレース値が、SLM8000 の元パルス値に対して 1/100 のスケールになっている
  ため、グラフ用計算式の中で ×100 を適用して元スケールに戻す。

【変更対象】
  - F列（θ機械位置→μm変換）：=(B{r}-$W$61)*$W$26  →  =(B{r}*100-$W$61)*$W$26
  - H列（Z機械位置→μm変換）：=(D{r}-$W$62)*$W$11  →  =(D{r}*100-$W$62)*$W$11

指令位置（A列θ指令位置、C列Z指令位置）はそのまま（×100 なし）。

【入力】 命名規則で最新の B（またはそれ以降）のファイル
【出力】 次のレター版（C, D, ...）

【実行前注意】
  Excel で該当ファイルを閉じておくこと。
"""

import sys
import os
import string
import re
from datetime import datetime
from openpyxl import load_workbook

DESKTOP = r"C:\Users\SEIGI-N13\Desktop"
BASE_NAME = "【巻線検証】SLM8000軌道_AISIN反映"


def find_latest_input():
    """最新の AISIN反映ファイルを探す（今日のラベルを降順で）"""
    td = datetime.now().strftime('%Y%m%d')
    for letter in reversed(string.ascii_uppercase):
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if os.path.exists(p):
            return p
    raise RuntimeError("AISIN反映ファイルが見つかりません")


def next_output_path():
    td = datetime.now().strftime('%Y%m%d')
    for letter in string.ascii_uppercase:
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if not os.path.exists(p):
            return p, letter
    raise RuntimeError(f"{td} は A-Z すべて使い切りました")


def patch_formulas(wb):
    ws = wb['NAKAM']
    # 置換パターン：\bB(\d+)-\$W\$61 → B{row}*100-$W$61
    #            ：\bD(\d+)-\$W\$62 → D{row}*100-$W$62
    # ただし既に *100 が付いている場合はスキップ（冪等）
    pat_B = re.compile(r'\bB(\d+)-\$W\$61\b')
    pat_D = re.compile(r'\bD(\d+)-\$W\$62\b')

    patched_F = 0
    patched_H = 0
    skipped = 0

    # F列、H列の全行を走査（NAKAM は最大 21808 行程度）
    last_row = ws.max_row
    print(f"  NAKAM max_row = {last_row}")

    for r in range(2, last_row + 1):
        for col, col_name in ((6, 'F'), (8, 'H')):
            cell = ws.cell(row=r, column=col)
            v = cell.value
            if v is None or not isinstance(v, str) or not v.startswith('='):
                continue
            # 既にパッチ済みかチェック（"B{r}*100" が含まれる場合）
            if f'B{r}*100' in v or f'D{r}*100' in v:
                skipped += 1
                continue

            new_v = v
            if col == 6:  # F列 θ
                new_v = pat_B.sub(lambda m: f'B{m.group(1)}*100-$W$61', new_v)
                if new_v != v:
                    cell.value = new_v
                    patched_F += 1
            elif col == 8:  # H列 Z
                new_v = pat_D.sub(lambda m: f'D{m.group(1)}*100-$W$62', new_v)
                if new_v != v:
                    cell.value = new_v
                    patched_H += 1

    print(f"  F列（θ機械位置→μm）パッチ: {patched_F} セル")
    print(f"  H列（Z機械位置→μm）パッチ: {patched_H} セル")
    print(f"  スキップ（既パッチ or 対象外）: {skipped} セル")
    return patched_F, patched_H


def main():
    input_path = find_latest_input()
    output_path, letter = next_output_path()
    print(f"入力: {input_path}")
    print(f"出力: {output_path}  （今日{letter}版）")

    print(f"\n読み込み中...")
    wb = load_workbook(input_path)
    print(f"シート: {wb.sheetnames}")

    print(f"\nF列、H列の数式に ×100 適用中...")
    pF, pH = patch_formulas(wb)

    # サンプル確認
    print(f"\n--- パッチ後サンプル ---")
    ws = wb['NAKAM']
    for r in (4, 5, 100, 1000):
        f = ws.cell(row=r, column=6).value
        h = ws.cell(row=r, column=8).value
        print(f"  row{r}: F={f}  H={h}")

    print(f"\n保存中...")
    try:
        wb.save(output_path)
        print(f"完了: {output_path}")
    except PermissionError:
        print(f"ERROR: 出力先が開かれています: {output_path}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()

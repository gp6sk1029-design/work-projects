# -*- coding: utf-8 -*-
"""
【パッチ用スクリプト】
既存データを保持したまま、ラベルの「送り」→「上下」のみ差し替える。
対象ファイル: 2976 LXLZ軌跡計算式_20230525_NEW.xlsx  シート: SLM8000方式_軌道計算

- データ行（入力・計算・参照）は一切触らない
- 書き換えるのは表題・注記・大見出し・中見出し・パラメータ説明のラベル文字列のみ
- 実行前に Excel でファイルを閉じておくこと
"""
import sys
import os
from openpyxl import load_workbook

PATH = r"C:\Users\SEIGI-N13\Desktop\2976 LXLZ軌跡計算式_20230525_NEW.xlsx"
SHEET = "SLM8000方式_軌道計算"

if not os.path.exists(PATH):
    print(f"ERROR: ファイルが見つかりません: {PATH}", file=sys.stderr)
    sys.exit(1)

print(f"読み込み: {PATH}")
wb = load_workbook(PATH)
if SHEET not in wb.sheetnames:
    print(f"ERROR: シートが見つかりません: {SHEET}", file=sys.stderr)
    sys.exit(1)
ws = wb[SHEET]

# 置き換え対象セルと新しい値
REPLACEMENTS = [
    # 行3（説明注記）
    ('A3', '【計算方式】SLM8000（NAKAM シート）ロジックを AISIN の三角関数式に移植。'
           'S/U 軸（振り）→ R·cosθ + √(L1²−(R·sinθ)²) − L1　（L1=LX!C4, R=LX!C7 を参照）。'
           'LZ 軸（上下）は同式で LZ!C4, LZ!C7 を参照。入力は Sysmac Studio の Pos [°]。'),
    # 行36 大見出し（左テーブル A36）
    ('A36', '◆ 左側：S軸（振り）＋LZ軸（上下） データ入力＆計算'),
    # 行36 大見出し（右テーブル L36）
    ('L36', '◆ 右側：U軸（振り）＋LZ軸（上下）※LZは左から自動参照'),
    # 行37 中見出し（左：E列 LZ入力）
    ('E37', 'LZ軸（上下）入力 [°]'),
    # 行37 中見出し（左：I列 LZ計算）
    ('I37', 'LZ軸→LZ座標 [mm]'),  # 変更なしだが念のため
    # 行37 中見出し（右：P列 LZ自動参照）
    ('P37', 'LZ軸（上下）※自動'),
    # 行37 中見出し（右：T列 LZ計算参照）
    ('T37', 'LZ軸→LZ座標※自動'),  # 変更なし
    # パラメータ説明（AB列）
    ('AB9',  'LZ軸（上下）の基準腕長'),
    ('AB10', 'LZ軸（上下）の振幅半径'),
]

count = 0
for addr, new_value in REPLACEMENTS:
    old_value = ws[addr].value
    if old_value != new_value:
        ws[addr].value = new_value
        count += 1
        print(f"  [UPD] {addr}: '{str(old_value)[:50]}...' → '{str(new_value)[:50]}...'")
    else:
        print(f"  [ -- ] {addr}: 変更なし")

# 保存
print(f"\n保存: {PATH}  （{count}セル更新）")
try:
    wb.save(PATH)
    print("完了")
except PermissionError:
    print("\nERROR: ファイルが他のプロセス（Excel）で開かれています。", file=sys.stderr)
    print("       Excel で該当ファイルを閉じてから再実行してください。", file=sys.stderr)
    sys.exit(1)

# 非回帰確認
print("\n--- 非回帰チェック ---")
wb2 = load_workbook(PATH)
ws2 = wb2[SHEET]
print(f"シート数: {len(wb2.sheetnames)}  新規シート最大行: {ws2.max_row}  最大列: {ws2.max_column}")
print(f"新規シート チャート数: {len(ws2._charts)}  (期待2)")

# データ入力列のサンプル確認（既入力値が残っていること）
print("\n--- 入力データ先頭行の非回帰（C/F/O列の値は保持されているか） ---")
for r in (39, 40, 41, 10038):
    c_val = ws2.cell(row=r, column=3).value
    f_val = ws2.cell(row=r, column=6).value
    o_val = ws2.cell(row=r, column=15).value
    print(f"  row{r}: C={c_val}  F={f_val}  O={o_val}")

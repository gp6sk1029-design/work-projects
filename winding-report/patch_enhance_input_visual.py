# -*- coding: utf-8 -*-
"""
【パッチ用スクリプト】入力エリアをさらに視覚的に強調

【目的】
  D版では色分けとコメントだけで分かりにくいため、以下を追加して
  ユーザーが迷わず入力場所を見つけられるようにする：

  1. 大きな矢印バナー「↓↓↓ ここにSysmacトレース値を貼り付け ↓↓↓」
  2. フリーズペイン（row 4, col 5）で入力ヘッダ常に表示
  3. 開いた時 A1 セルにジャンプ（スクロール位置リセット）
  4. 入力エリアに太い色付き枠（囲み枠）
  5. 初回サンプル1行（行4）に薄くプレースホルダー値 → 上書きイメージ
  6. 入力エリアの列幅を広めに
"""

import sys
import os
import string
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

DESKTOP = r"C:\Users\SEIGI-N13\Desktop"
BASE_NAME = "【巻線検証】SLM8000軌道_AISIN反映"


def find_latest_input():
    td = datetime.now().strftime('%Y%m%d')
    for letter in reversed(string.ascii_uppercase):
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if os.path.exists(p):
            return p
    raise RuntimeError("ファイルが見つかりません")


def next_output_path():
    td = datetime.now().strftime('%Y%m%d')
    for letter in string.ascii_uppercase:
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if not os.path.exists(p):
            return p, letter
    raise RuntimeError("使い切り")


# スタイル
FILL_BANNER    = PatternFill('solid', fgColor='FFC000')  # オレンジ（警告色で目立つ）
FILL_CMD       = PatternFill('solid', fgColor='DEEBF7')
FILL_ACT       = PatternFill('solid', fgColor='FCE4D6')
FILL_HEADER    = PatternFill('solid', fgColor='1F4E78')
FONT_BANNER    = Font(name='游ゴシック', size=14, bold=True, color='000000')
FONT_TITLE     = Font(name='游ゴシック', size=12, bold=True, color='FFFFFF')
FONT_HEADER    = Font(name='游ゴシック', size=10, bold=True, color='FFFFFF')
FONT_SIGNAL    = Font(name='游ゴシック', size=9, color='1F4E78', italic=True)
FONT_PLACEHOLDER = Font(name='游ゴシック', size=9, color='BFBFBF', italic=True)
ALIGN_C        = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 太い色付き枠（入力エリア囲み用）
BORDER_INPUT = Border(
    left=Side(style='medium', color='C00000'),
    right=Side(style='medium', color='C00000'),
    top=Side(style='medium', color='C00000'),
    bottom=Side(style='medium', color='C00000'),
)


def apply_patch(wb):
    ws = wb['NAKAM']
    print(f"  NAKAM max_row={ws.max_row}")

    # ─────────────────────────
    # 既存の結合セルで row 1-3, col 1-4 にかかるものを解除
    # ─────────────────────────
    for mr in list(ws.merged_cells.ranges):
        if (mr.min_row <= 3 and mr.max_row >= 1 and
            mr.min_col <= 4 and mr.max_col >= 1):
            ws.unmerge_cells(str(mr))

    # ─────────────────────────
    # 1. 列幅を広めに（col 1-4）
    # ─────────────────────────
    for c, w in [(1, 22), (2, 22), (3, 22), (4, 22)]:
        ws.column_dimensions[chr(64 + c)].width = w

    # ─────────────────────────
    # 2. row 1: 大きな橙色バナー「↓↓↓ Sysmacトレース値をここに貼り付け ↓↓↓」
    # ─────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    cell = ws.cell(row=1, column=1)
    cell.value = '↓↓↓ Sysmac Studio データトレース値をここに貼り付け ↓↓↓'
    cell.font = FONT_BANNER
    cell.fill = FILL_BANNER
    cell.alignment = ALIGN_C
    ws.row_dimensions[1].height = 34

    # ─────────────────────────
    # 3. row 2: 軸区分（θ / Z）
    # ─────────────────────────
    # row 2 col 1-2 = θ（結合）、col 3-4 = Z（結合）
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
    c_theta = ws.cell(row=2, column=1)
    c_theta.value = 'θ 軸（首振り）'
    c_theta.font = FONT_TITLE
    c_theta.fill = FILL_HEADER
    c_theta.alignment = ALIGN_C
    c_z = ws.cell(row=2, column=3)
    c_z.value = 'Z 軸（上下）'
    c_z.font = FONT_TITLE
    c_z.fill = FILL_HEADER
    c_z.alignment = ALIGN_C
    ws.row_dimensions[2].height = 24

    # ─────────────────────────
    # 4. row 3: 信号名ヘッダ（Cmd/Act 各列）
    # ─────────────────────────
    sig_names = [
        (1, '指令位置\nMC_Axis_θ.Cmd.Pos', FILL_CMD),
        (2, '機械位置 (×100)\nMC_Axis_θ.Act.Pos', FILL_ACT),
        (3, '指令位置\nMC_Axis_Z.Cmd.Pos', FILL_CMD),
        (4, '機械位置 (×100)\nMC_Axis_Z.Act.Pos', FILL_ACT),
    ]
    for col, name, fill in sig_names:
        cell = ws.cell(row=3, column=col)
        cell.value = name
        cell.font = FONT_SIGNAL
        cell.fill = fill
        cell.alignment = ALIGN_C
    ws.row_dimensions[3].height = 44

    # ─────────────────────────
    # 5. row 4: プレースホルダ（灰色斜体で「ここに値を貼り付け」）
    # ─────────────────────────
    placeholders = [
        (1, '← ここに貼り付け'),
        (2, '← ここに貼り付け'),
        (3, '← ここに貼り付け'),
        (4, '← ここに貼り付け'),
    ]
    for col, text in placeholders:
        cell = ws.cell(row=4, column=col)
        if cell.value is None:  # 既に値があればスキップ
            cell.value = text
            cell.font = FONT_PLACEHOLDER
            cell.alignment = ALIGN_C

    # ─────────────────────────
    # 6. 入力エリア全体に太い赤枠（row 1-3 タイトル + row 4-5042 入力）
    # ─────────────────────────
    # 上辺・下辺・左辺・右辺を囲む
    last_input_row = 5042  # F列数式が有効な範囲まで（前パッチで確認済）

    # 上辺（row 1 の上端）
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cur = cell.border
        cell.border = Border(
            top=Side(style='medium', color='C00000'),
            bottom=cur.bottom if cur else None,
            left=Side(style='medium', color='C00000') if c == 1 else (cur.left if cur else None),
            right=Side(style='medium', color='C00000') if c == 4 else (cur.right if cur else None),
        )
    # 下辺（last_input_row の下端）※多すぎると重いので、最後のセル数行だけ枠
    # 実用的には「入力可能エリア」の明示で row 1-3 + row 4 だけ強調枠
    # （row 4 〜 last_input_row の長い枠は Excel が描画しても分かりにくいためスキップ）

    # row 4 に太い上下左右枠を付けて「最初の行」を強調
    for c in range(1, 5):
        cell = ws.cell(row=4, column=c)
        cell.border = Border(
            top=Side(style='medium', color='C00000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='medium', color='C00000') if c == 1 else Side(style='thin', color='BFBFBF'),
            right=Side(style='medium', color='C00000') if c == 4 else Side(style='thin', color='BFBFBF'),
        )

    # ─────────────────────────
    # 7. フリーズペイン（row 4 より上 + col 5 より左を固定）
    # ─────────────────────────
    ws.freeze_panes = 'E4'
    print(f"  freeze_panes = E4（row 1-3 + col A-D を固定表示）")

    # ─────────────────────────
    # 8. シートを開いたときのアクティブセル・スクロール位置を A1 に
    # ─────────────────────────
    ws.sheet_view.topLeftCell = 'A1'
    ws.sheet_view.selection[0].activeCell = 'A4'
    ws.sheet_view.selection[0].sqref = 'A4'

    # NAKAM を最初に表示するシートに
    wb.active = wb.sheetnames.index('NAKAM')

    # コメント再設定（既存があるならそのまま）
    for col, desc in [
        (1, 'MC_Axis_θ.Cmd.Pos をここから下に貼り付け。\n（×100 補正なし）'),
        (2, 'MC_Axis_θ.Act.Pos をここから下に貼り付け。\n（×100 補正：グラフ計算で自動適用）'),
        (3, 'MC_Axis_Z.Cmd.Pos をここから下に貼り付け。\n（×100 補正なし）'),
        (4, 'MC_Axis_Z.Act.Pos をここから下に貼り付け。\n（×100 補正：グラフ計算で自動適用）'),
    ]:
        cell = ws.cell(row=4, column=col)
        cell.comment = Comment(desc, 'AISIN反映版')

    print(f"  視覚的強調（バナー・フリーズペイン・枠・ジャンプ位置）適用完了")


def main():
    input_path = find_latest_input()
    output_path, letter = next_output_path()
    print(f"入力: {input_path}")
    print(f"出力: {output_path}  （今日{letter}版）")

    wb = load_workbook(input_path)
    print(f"\n--- パッチ適用中 ---")
    apply_patch(wb)

    print(f"\n--- 保存 ---")
    try:
        wb.save(output_path)
        print(f"完了: {output_path}")
    except PermissionError:
        print(f"ERROR: 出力先が開かれています: {output_path}", file=sys.stderr)
        sys.exit(1)

    # 検証
    wb2 = load_workbook(output_path)
    ws = wb2['NAKAM']
    print(f"\n--- 検証 ---")
    print(f"  freeze_panes: {ws.freeze_panes}")
    print(f"  sheet active: {wb2.active.title}")
    print(f"  row 1 col 1: {ws.cell(row=1, column=1).value}")
    print(f"  row 2 col 1/3: {ws.cell(row=2, column=1).value!r} | {ws.cell(row=2, column=3).value!r}")
    print(f"  row 3 col 1-4:")
    for c in range(1, 5):
        print(f"    col{c}: {ws.cell(row=3, column=c).value!r}")
    print(f"  row 4 col 1-4:")
    for c in range(1, 5):
        print(f"    col{c}: {ws.cell(row=4, column=c).value!r}")
    # 数式確認（row 4 col 5-8）
    print(f"  row 4 col 5-8（数式）:")
    for c in range(5, 9):
        print(f"    col{c}: {ws.cell(row=4, column=c).value}")


if __name__ == '__main__':
    main()

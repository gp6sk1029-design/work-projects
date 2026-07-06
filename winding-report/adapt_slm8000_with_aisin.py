# -*- coding: utf-8 -*-
"""
SLM8000 ファイルを土台に、AISIN の製品（ワーク）データを反映させて新ファイル生成

【目的】
  SLM8000 の計算ロジック・カム設定・グラフ構造を維持したまま、
  ワーク大きさ（NAKAM シート col 16-18）を AISIN の製品輪郭で置換する。

【入力】
  - SLM8000 元ファイル（.xls）  ：計算ロジック・グラフ構造の土台
  - AISIN 元ファイル（.xlsx）   ：製品輪郭データ（等加速度20230825!F6:G36）

【出力】
  - 新ファイル（.xlsx） ：命名規則 {ベース名}_{YYYYMMDD}{A-Z}.xlsx

【処理フロー】
  1. Excel COM で SLM8000 .xls → .xlsx に変換（数式・グラフ保持）
  2. AISIN の製品輪郭 31 点を読み取り
  3. NAKAM シートの ワーク大きさ（col 16-18）を AISIN データで置換
  4. mm → μm 変換（×1000）
  5. 既存の SLM8000 データ行（row 2-22 程度）をクリアして AISIN データ（31 行）を書き込み
"""

import sys
import os
import string
from datetime import datetime
from openpyxl import load_workbook
import win32com.client as win32

# ════════════════════════════════════════
# パス設定
# ════════════════════════════════════════
DESKTOP = r"C:\Users\SEIGI-N13\Desktop"
SLM_XLS = os.path.join(DESKTOP, "【巻線検証】SLM8000軌道_H2X ZS L 400rpm.xls")
AISIN_XLSX = os.path.join(DESKTOP, "2976 LXLZ軌跡計算式_20230525.xlsx")
BASE_NAME = "【巻線検証】SLM8000軌道_AISIN反映"


def next_output_path():
    """命名規則 {ベース名}_{YYYYMMDD}{A-Z}.xlsx に従う次の空きパス"""
    td = datetime.now().strftime('%Y%m%d')
    for letter in string.ascii_uppercase:
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if not os.path.exists(p):
            return p, letter
    raise RuntimeError(f"{td} は A-Z すべて使い切りました")


def xls_to_xlsx_via_excel(xls_path, xlsx_path):
    """Excel COM で .xls → .xlsx 変換（数式・グラフ保持）"""
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(xls_path))
        # FileFormat=51 = xlOpenXMLWorkbook (.xlsx)
        wb.SaveAs(os.path.abspath(xlsx_path), FileFormat=51)
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()


def read_aisin_outline(path):
    """AISIN 等加速度20230825!F6:G36 を (X_mm, Y_mm) 点列で取得"""
    wb = load_workbook(path, data_only=True)
    ws = wb['等加速度20230825']
    points = []
    for r in range(6, 37):  # F6:G36 = 31 行
        x = ws.cell(row=r, column=6).value   # F列
        y = ws.cell(row=r, column=7).value   # G列
        if x is None or y is None:
            continue
        points.append((x, y))
    return points


def apply_aisin_to_slm(xlsx_path, aisin_points):
    """
    NAKAM シートの ワーク大きさ（col 16-18）に AISIN データを反映
    - col 16: ラベル（P1, P2, ..., AISIN製品輪郭 と表示）
    - col 17: X座標 [um] = AISIN_X_mm × 1000
    - col 18: Y座標 [um] = AISIN_Y_mm × 1000
    """
    wb = load_workbook(xlsx_path)
    if 'NAKAM' not in wb.sheetnames:
        print(f"ERROR: NAKAM シートが見つかりません。シート一覧: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb['NAKAM']

    # 対象範囲にかかる結合セルをすべて解除（MergedCellは書き込み不可）
    merged_to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        # row 2〜50 と col 16-18 のどこかに重なっていれば解除対象
        if (mr.min_row <= 50 and mr.max_row >= 2 and
            mr.min_col <= 18 and mr.max_col >= 16):
            merged_to_unmerge.append(str(mr))
    for mr_str in merged_to_unmerge:
        ws.unmerge_cells(mr_str)
    if merged_to_unmerge:
        print(f"  結合セル {len(merged_to_unmerge)} 箇所を解除: {merged_to_unmerge[:5]}...")

    # 既存のワーク大きさデータを row 2〜50 までクリア（安全のため広めに）
    cleared = 0
    for r in range(2, 51):
        for c in (16, 17, 18):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None
                cleared += 1
    print(f"  既存ワーク大きさ {cleared} セルをクリア")

    # タイトル行（row 1）は既存のまま（'ワーク大きさ'）

    # AISIN データ書き込み（row 2 から）
    for i, (x_mm, y_mm) in enumerate(aisin_points):
        r = 2 + i
        # ラベル（どの点かが分かる命名）
        label = f'P{i+1:02d}'
        if i == 0:
            label = 'AISIN起点'
        elif i == len(aisin_points) - 1:
            label = 'AISIN終点'
        ws.cell(row=r, column=16).value = label
        # mm → um 変換
        ws.cell(row=r, column=17).value = round(x_mm * 1000, 2)  # X [um]
        ws.cell(row=r, column=18).value = round(y_mm * 1000, 2)  # Y [um]

    print(f"  AISIN製品輪郭 {len(aisin_points)} 点を書き込み（row 2 〜 {2 + len(aisin_points) - 1}）")
    print(f"  X範囲: {min(p[0] for p in aisin_points):.3f} 〜 {max(p[0] for p in aisin_points):.3f} mm")
    print(f"  Y範囲: {min(p[1] for p in aisin_points):.3f} 〜 {max(p[1] for p in aisin_points):.3f} mm")

    # ─────────────────────────
    # チャート更新：AISIN データが収まるように軸範囲拡大＆系列範囲延長
    # ─────────────────────────
    last_data_row = 2 + len(aisin_points) - 1   # row 32（31点の場合）
    # AISIN の絶対値最大を取得してマージンを持たせた軸範囲を算出
    x_max_um = max(abs(p[0]) for p in aisin_points) * 1000   # μm
    y_abs_max_um = max(abs(p[1]) for p in aisin_points) * 1000
    # 10%マージン追加 → 500単位で切り上げ
    def round_up(v, step=500):
        return int((v * 1.1 + step - 1) // step * step)
    x_range = round_up(x_max_um)   # 例: 4640 × 1.1 = 5104 → 5500
    y_range = round_up(y_abs_max_um)  # 例: 11350 × 1.1 = 12485 → 12500

    for ch in ws._charts:
        # 軸範囲を拡大（対称）
        ch.x_axis.scaling.min = -x_range
        ch.x_axis.scaling.max =  x_range
        ch.y_axis.scaling.min = -y_range
        ch.y_axis.scaling.max =  y_range
        print(f"  チャート軸更新: X=±{x_range}  Y=±{y_range}")

        # 系列2（ワーク = SLM8000 オリジナルでは NAKAM!$Q$2:$Q$22 → 21行）
        # AISIN は 31点 なので row 2 〜 last_data_row へ拡張
        # ただし系列の内容が "ワーク" の参照（Q列/R列）であるときのみ更新
        for s in ch.series:
            try:
                xf = s.xVal.numRef.f if s.xVal and s.xVal.numRef else ''
                yf = s.yVal.numRef.f if s.yVal and s.yVal.numRef else ''
                # "$Q$" と "$R$" を参照している系列を特定（ワーク輪郭）
                if '$Q$' in xf and '$R$' in yf and 'NAKAM' in xf:
                    new_xf = f"NAKAM!$Q$2:$Q${last_data_row}"
                    new_yf = f"NAKAM!$R$2:$R${last_data_row}"
                    s.xVal.numRef.f = new_xf
                    s.yVal.numRef.f = new_yf
                    print(f"  ワーク系列参照を更新: x={new_xf}  y={new_yf}")
            except Exception as e:
                print(f"    (series 更新スキップ: {e})")

    # 保存
    wb.save(xlsx_path)


def main():
    if not os.path.exists(SLM_XLS):
        print(f"ERROR: SLM8000 ファイルが見つかりません: {SLM_XLS}", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(AISIN_XLSX):
        print(f"ERROR: AISIN ファイルが見つかりません: {AISIN_XLSX}", file=sys.stderr)
        sys.exit(1)

    # 出力パス決定
    output_path, letter = next_output_path()
    print(f"入力 SLM:   {SLM_XLS}")
    print(f"入力 AISIN: {AISIN_XLSX}")
    print(f"出力:       {output_path}  （今日{letter}版）")

    # Step 1: SLM8000 xls → xlsx 変換
    print(f"\n[1/3] Excel COM で .xls → .xlsx 変換中...")
    try:
        xls_to_xlsx_via_excel(SLM_XLS, output_path)
        print(f"       変換完了（{os.path.getsize(output_path):,} bytes）")
    except Exception as e:
        print(f"       ERROR: 変換失敗 {e}", file=sys.stderr)
        sys.exit(1)

    # Step 2: AISIN 輪郭データ取得
    print(f"\n[2/3] AISIN 製品輪郭データを読み込み中...")
    aisin_points = read_aisin_outline(AISIN_XLSX)
    print(f"       取得: {len(aisin_points)} 点")
    print(f"       先頭3点: {aisin_points[:3]}")
    print(f"       末尾3点: {aisin_points[-3:]}")

    # Step 3: SLM8000 xlsx に AISIN データを反映
    print(f"\n[3/3] NAKAM シートの ワーク大きさ を AISIN データで置換中...")
    apply_aisin_to_slm(output_path, aisin_points)

    # 検証
    print(f"\n--- 検証 ---")
    wb = load_workbook(output_path)
    print(f"  シート一覧: {wb.sheetnames}")
    nakam = wb['NAKAM']
    print(f"  NAKAM サイズ: {nakam.max_row} 行 × {nakam.max_column} 列")
    print(f"  NAKAM charts: {len(nakam._charts)}")
    print(f"  ワーク大きさ先頭5行:")
    for r in range(2, 7):
        label = nakam.cell(row=r, column=16).value
        x = nakam.cell(row=r, column=17).value
        y = nakam.cell(row=r, column=18).value
        print(f"    row{r}: label={label}  X={x}  Y={y}")
    print(f"  ワーク大きさ末尾3行:")
    for r in range(30, 33):
        label = nakam.cell(row=r, column=16).value
        x = nakam.cell(row=r, column=17).value
        y = nakam.cell(row=r, column=18).value
        print(f"    row{r}: label={label}  X={x}  Y={y}")

    print(f"\n完了: {output_path}")


if __name__ == '__main__':
    main()

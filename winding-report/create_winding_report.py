# -*- coding: utf-8 -*-
"""
EOP巻線機 MGワイヤ傷 改善報告書

【恒久的対策の核心】
  openpyxlで結合セル（merge_cells）を使うと、Excelの「行高自動調整」が
  一切機能しない。そのため、テキスト量から行高を数式で算出する
  calc_h() 関数を実装し、全テキスト行に適用する。

  calc_h(text, col_span, font_size) の計算式:
    1. 列幅(pt) = col_span × 8.5[unit] × 5.5[pt/unit]
    2. 1行に収まる文字数 = 列幅(pt) ÷ (font_size × 0.75)
    3. 折り返し行数 = Σ ceil(各行文字数 ÷ 1行文字数)
    4. 行高(pt) = 折り返し行数 × (font_size × 1.6) + 上下パディング8pt
"""
import math
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
ws = wb.active
ws.title = '改善報告書'

IMG = 'C:/Users/SEIGI-N13/Desktop/temp_images'
YG  = '游ゴシック'
COL_W = 8.5   # 標準列幅（Excelの列幅単位）

# ════════════════════════════════════════
# ★ 恒久的解決策: 行高自動計算関数
# ════════════════════════════════════════
def calc_h(text, col_span, font_size=11, min_h=22):
    """
    結合セルに入るテキストから適切な行高(pt)を計算する。

    Args:
        text     : セルに書くテキスト（\n含む可）
        col_span : 結合する列数
        font_size: フォントサイズ(pt)
        min_h    : 最低行高(pt)

    Returns:
        行高(pt) ※ Excel row_dimensions.height に設定する値
    """
    if not text:
        return min_h

    # 列幅をポイント換算: 1 Excel列幅unit ≈ 5.5pt（游ゴシック基準）
    col_width_pt  = col_span * COL_W * 5.5

    # 日本語1文字の幅(pt): 游ゴシック は正方形に近い → font_size × 0.75
    char_width_pt = font_size * 0.75

    # 1行に入る最大文字数
    chars_per_line = max(1, math.floor(col_width_pt / char_width_pt))

    # 各改行ごとに折り返し行数を積算
    total_lines = 0
    for line in text.split('\n'):
        total_lines += max(1, math.ceil(len(line) / chars_per_line)) if line else 1

    # 1行の高さ = font_size × 1.6 + パディング
    return max(min_h, math.ceil(total_lines * font_size * 1.6) + 8)

# ════════════════════════════════════════
# スタイル定義（全て遊ゴシック）
# ════════════════════════════════════════
def F(**kw): return Font(name=YG, **kw)

f_title   = F(size=16, bold=True,  color='FFFFFF')
f_section = F(size=14, bold=True,  color='1F3864')
f_sub     = F(size=12, bold=True)
f_body    = F(size=11)
f_note    = F(size=10, color='595959', italic=True)
f_label_r = F(size=11, bold=True,  color='C00000')
f_label_g = F(size=11, bold=True,  color='00692A')
f_white   = F(size=11, bold=True,  color='FFFFFF')
f_white_s = F(size=10, color='AAAAAA')
f_sch     = F(size=11, bold=True)
f_anno_r  = F(size=10, bold=True,  color='C00000')
f_anno_b  = F(size=10, bold=True,  color='1F3864')
f_anno_g  = F(size=10, bold=True,  color='006400')
f_footer  = F(size=11, color='999999')
f_flow    = F(size=12, bold=True,  color='1F3864')
f_warn    = F(size=11, bold=True,  color='7F5200')

def P(c): return PatternFill('solid', fgColor=c)
p_title   = P('1F3864'); p_section = P('D9E2F3')
p_red_l   = P('FFF0F0'); p_green_l = P('EDF7ED')
p_yellow  = P('FFFACD'); p_gray    = P('F2F2F2')
p_sch1    = P('E8F0FE'); p_sch2    = P('C6DBEF')
p_anno_r  = P('FFECEC'); p_anno_g  = P('EBFAEB'); p_anno_b  = P('E3EEFF')

ac = Alignment(horizontal='center', vertical='center', wrap_text=True)
al = Alignment(horizontal='left',   vertical='center', wrap_text=True)
at = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
ar = Alignment(horizontal='right',  vertical='center', wrap_text=True)

thin  = Side(style='thin', color='BFBFBF')
B_all = Border(left=thin, right=thin, top=thin, bottom=thin)

# ════════════════════════════════════════
# ヘルパー関数
# ════════════════════════════════════════
def sty(cell, font=f_body, fill=None, align=al):
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = align

def rh(row, h):
    ws.row_dimensions[row].height = h

def mw(row, cs, ce, text, font=f_body, fill=None, align=al, h=None):
    """
    セル結合して書き込み。
    h を指定しない場合は calc_h() で自動計算する（★恒久対策）
    """
    c1, c2 = get_column_letter(cs), get_column_letter(ce)
    ws.merge_cells(f'{c1}{row}:{c2}{row}')
    c = ws.cell(row=row, column=cs, value=text)
    sty(c, font=font, fill=fill, align=align)

    if h is None:
        # ★ テキスト行は必ず自動計算（結合列数で計算）
        auto = calc_h(text or '', ce - cs + 1, font.size or 11)
        rh(row, auto)
    else:
        rh(row, h)
    return c

def section_bar(row, text):
    mw(row, 2, 18, text, font=f_section, fill=p_section, align=al, h=30)

def cap(row, specs, h=20):
    """複数キャプションを一行に並べる。specs=[(col_s,col_e,text,font,fill), ...]"""
    rh(row, h)
    for cs, ce, text, font, fill in specs:
        c1, c2 = get_column_letter(cs), get_column_letter(ce)
        ws.merge_cells(f'{c1}{row}:{c2}{row}')
        c = ws.cell(row=row, column=cs, value=text)
        sty(c, font=font, fill=fill, align=ac)

def ann(row, specs, h=20):
    """複数注釈を一行に並べる（capと同じ構造）"""
    cap(row, specs, h)

def img_row(row, h_cm):
    """画像専用行: 1cm=28.35pt + 余裕15pt"""
    rh(row, h_cm * 28.35 + 15)

def add_img(path, col, row, w_cm, h_cm):
    fpath = os.path.join(IMG, path)
    if not os.path.exists(fpath):
        print(f'  ⚠ 画像なし: {path}'); return
    img = Image(fpath)
    img.width  = w_cm * 37.8
    img.height = h_cm * 37.8
    ws.add_image(img, f'{get_column_letter(col)}{row}')

def brd(row, col_s, col_e):
    for col in range(col_s, col_e+1):
        ws.cell(row=row, column=col).border = B_all

# 列幅設定
ws.column_dimensions['A'].width = 3
for col in range(2, 19):
    ws.column_dimensions[get_column_letter(col)].width = COL_W

# ════════════════════════════════════════════════════════════
R = 1   # 行カーソル

# ──────────────────────────────────────
# ① 表紙ヘッダー
# ──────────────────────────────────────
for r in range(R, R+4):
    for col in range(2, 19): ws.cell(row=r, column=col).fill = p_title

mw(R, 2, 18, '',                                       fill=p_title, h=8);   R+=1
mw(R, 2, 18, 'EOP巻線機 MGワイヤ傷 改善報告書',
   font=f_title, fill=p_title, align=ac, h=42);                               R+=1
mw(R, 2, 10, '報告日: 2026年4月9日',
   font=f_white, fill=p_title, align=al, h=28)
mw(R, 11, 18, '新機構（コア固定方式）導入に伴う問題と改善の報告',
   font=f_white_s, fill=p_title, align=ar, h=28);                             R+=1
mw(R, 2, 18, '',                                       fill=p_title, h=6);   R+=1
rh(R, 10); R+=1

# ──────────────────────────────────────
# ② 報告サマリー
# ──────────────────────────────────────
section_bar(R, '■ 報告サマリー'); R+=1

mw(R, 2, 18,
   '本報告書は、EOP巻線機における新機構（コア固定方式）導入時に発生したMGワイヤの傷・ピンホール問題について、'
   '原因分析・実施した対策・改善結果をまとめたものです。',
   font=f_body); R+=1   # ← h省略 → calc_h()が自動計算

# ステータス表ヘッダー
mw(R, 2, 3,   '項目',       font=f_sub, fill=p_gray, align=ac, h=26)
mw(R, 4, 13,  '内容',       font=f_sub, fill=p_gray, align=ac, h=26)
mw(R, 14, 16, 'ステータス', font=f_sub, fill=p_gray, align=ac, h=26)
brd(R, 2, 16); R+=1

for item, desc, status, bg, sf in [
    ('課題1', 'フック外れ → スピード・高さ調整、電線渡り溝の追加により解決', '対策済み', p_green_l, f_label_g),
    ('課題2', '下渡り傷 → 治具バフ研磨、コア押上げリングのバリ除去により解決', '対策済み', p_green_l, f_label_g),
    ('巻線部', 'S軸(左)表面にピンホールあり → ゲイン調整により対応中', '調整中', p_yellow, f_label_r),
    ('巻線整列', 'プログラム変更なし。目視にて良好な整列を確認', '良好', p_green_l, f_label_g),
]:
    # 行高は最も行数が多いセル(内容列=10列)で決まる
    h = calc_h(desc, 10, 11)
    mw(R, 2, 3,   item,   font=f_body, fill=bg, align=ac, h=h)
    mw(R, 4, 13,  desc,   font=f_body, fill=bg, align=al, h=h)
    mw(R, 14, 16, status, font=sf,     fill=bg, align=ac, h=h)
    brd(R, 2, 16); R+=1

rh(R, 10); R+=1

# ──────────────────────────────────────
# ③ 課題1: フック外れ
# ──────────────────────────────────────
section_bar(R, '■ 課題1: フック外れ ── 対策済み'); R+=1

mw(R, 2, 18, '【問題】', font=f_sub, h=26); R+=1
mw(R, 2, 18,
   '巻線工程において、電線をフックに掛ける際にフック外れが発生しておりました。'
   'フック外れが起きると電線の経路が乱れ、巻線品質に影響を及ぼします。',
   font=f_body); R+=1   # 自動計算

mw(R, 2, 18, '【対策内容】', font=f_sub, h=26); R+=1
mw(R, 2, 18,
   '① フックのスピードを低速に調整いたしました。\n'
   '② フックの高さを調整いたしました（1.0mmシムを1枚追加）。\n'
   '③ 電線を安定的に渡らせるための溝を治具に追加し、確実な渡りを実現いたしました。',
   font=f_body); R+=1   # 自動計算（\nを考慮して3行以上で計算）

cap(R, [
    (2,  9,  '▼ 電線渡り溝を追加した治具',   f_note, p_anno_b),
    (10, 17, '▼ 下渡り治具（改善後の状態）', f_note, p_anno_b),
]); R+=1

img_row(R, 5)
add_img('image17.png', 2,  R, w_cm=7, h_cm=5)
add_img('image13.png', 10, R, w_cm=7, h_cm=5); R+=1

ann(R, [
    (2,  9,  '↑ 新設した電線渡り溝（安定渡りを実現）', f_anno_b, p_anno_b),
    (10, 17, '↑ 溝加工部：ガイド溝が追加された状態',  f_anno_b, p_anno_b),
]); R+=1
rh(R, 10); R+=1

# ──────────────────────────────────────
# ④ 課題2: 下渡り傷
# ──────────────────────────────────────
section_bar(R, '■ 課題2: 下渡り傷（ワイヤ傷・ピンホール） ── 対策済み'); R+=1

# ④-a 改善前
mw(R, 2, 8, '【改善前の状態】', font=f_sub, fill=p_red_l, h=26)
mw(R, 9, 11, '⚠ 改善前',       font=f_label_r, fill=p_red_l, align=ac, h=26); R+=1

mw(R, 2, 18,
   '下渡り部においてワイヤ表面に傷およびピンホールが発生しておりました。'
   '電線が治具と接触する部分で、表面の被膜が損傷している状態が確認されました。',
   font=f_body); R+=1   # 自動計算

cap(R, [
    (2,  5,  '▼ 下渡り部の傷（全体）',   f_note, p_anno_r),
    (6,  9,  '▼ 下渡り部の傷（拡大）',   f_note, p_anno_r),
    (10, 13, '▼ ワイヤ表面のバリ・傷',   f_note, p_anno_r),
    (14, 17, '▼ ピンホール発生箇所',     f_note, p_anno_r),
]); R+=1

img_row(R, 5.5)
add_img('image7.jpeg',  2,  R, w_cm=7, h_cm=5.5)
add_img('image8.jpg',   6,  R, w_cm=7, h_cm=5.5)
add_img('image9.jpeg',  10, R, w_cm=7, h_cm=5.5)
add_img('image10.jpeg', 14, R, w_cm=7, h_cm=5.5); R+=1

ann(R, [
    (2,  5,  '↑ 傷の発生箇所',          f_anno_r, p_anno_r),
    (6,  9,  '↑ 被膜の損傷（拡大）',    f_anno_r, p_anno_r),
    (10, 13, '↑ バリによる線条痕',      f_anno_r, p_anno_r),
    (14, 17, '↑ 赤色部 = ピンホール',   f_anno_r, p_anno_r),
]); R+=1
rh(R, 10); R+=1

# ④-b 原因分析
mw(R, 2, 18, '【原因分析と対策】', font=f_sub, fill=p_section, h=26); R+=1
mw(R, 2, 18,
   '原因1: 下渡り治具の表面粗さにより、電線被膜に傷が発生\n'
   '　→ 対策: 治具表面のバフ研磨を実施いたしました。\n'
   '原因2: コア押上げリングの摩耗によりバリが発生し、電線に接触\n'
   '　→ 対策: リングのバフ研磨および清掃を実施いたしました。',
   font=f_body); R+=1   # 自動計算

cap(R, [
    (2,  7,  '▼ 巻線機 断面図（CAD）',     f_note, p_anno_b),
    (8,  12, '▼ コア押上げリング（部品）', f_note, p_anno_b),
    (13, 17, '▼ リング組み付け状態',       f_note, p_anno_b),
]); R+=1

img_row(R, 6)
add_img('image16.png', 2,  R, w_cm=11, h_cm=6)
add_img('image15.png', 8,  R, w_cm=8,  h_cm=6)
add_img('image14.png', 13, R, w_cm=8,  h_cm=6); R+=1

ann(R, [
    (2,  4,  '↑ ワイヤ経路',            f_anno_b, p_anno_b),
    (5,  7,  '↑ リング位置',            f_anno_b, p_anno_b),
    (8,  12, '↑ バリ発生箇所を研磨',   f_anno_r, p_anno_r),
    (13, 17, '↑ 組付け後の確認状態',   f_anno_b, p_anno_b),
]); R+=1
rh(R, 10); R+=1

# ④-c 改善後
mw(R, 2, 8,  '【改善後の結果】', font=f_sub, fill=p_green_l, h=26)
mw(R, 9, 11, '✓ 改善後',        font=f_label_g, fill=p_green_l, align=ac, h=26); R+=1

mw(R, 2, 18, '▼ 左軸（S軸） ── 改善後の下渡り状態',
   font=f_sub, fill=p_green_l, h=26); R+=1

cap(R, [
    (2,  5,  '拡大写真①', f_note, p_anno_g),
    (6,  9,  '拡大写真②', f_note, p_anno_g),
    (10, 13, '比較画像',   f_note, p_anno_g),
    (14, 17, '全体写真',   f_note, p_anno_g),
]); R+=1

img_row(R, 5.5)
add_img('image1.jpeg',  2,  R, w_cm=7, h_cm=5.5)
add_img('image2.jpeg',  6,  R, w_cm=7, h_cm=5.5)
add_img('image3.png',   10, R, w_cm=7, h_cm=5.5)
add_img('image11.jpeg', 14, R, w_cm=7, h_cm=5.5); R+=1

ann(R, [
    (2,  5,  '↑ 傷なし・状態良好', f_anno_g, p_anno_g),
    (6,  9,  '↑ 被膜損傷なし',     f_anno_g, p_anno_g),
    (10, 13, '↑ 改善前後の比較',   f_anno_b, p_anno_b),
    (14, 17, '↑ S軸全体良好',      f_anno_g, p_anno_g),
]); R+=1

mw(R, 2, 18,
   '※ 下渡り部の傷は確認されず、安定した状態で巻線が行われていることを確認いたしました。',
   font=f_note, align=al); R+=1
rh(R, 8); R+=1

mw(R, 2, 18, '▼ 右軸（U軸） ── 改善後の下渡り状態',
   font=f_sub, fill=p_green_l, h=26); R+=1

cap(R, [
    (2,  5,  '拡大写真①',  f_note, p_anno_g),
    (6,  9,  '拡大写真②',  f_note, p_anno_g),
    (10, 13, '全体写真①', f_note, p_anno_g),
    (14, 17, '全体写真②', f_note, p_anno_g),
]); R+=1

img_row(R, 5.5)
add_img('image4.jpeg',  2,  R, w_cm=7, h_cm=5.5)
add_img('image5.jpeg',  6,  R, w_cm=7, h_cm=5.5)
add_img('image6.jpg',   10, R, w_cm=7, h_cm=5.5)
add_img('image12.jpeg', 14, R, w_cm=7, h_cm=5.5); R+=1

ann(R, [
    (2,  5,  '↑ 傷なし・状態良好', f_anno_g, p_anno_g),
    (6,  9,  '↑ ピンホールなし',   f_anno_g, p_anno_g),
    (10, 13, '↑ U軸全体良好',      f_anno_g, p_anno_g),
    (14, 17, '↑ 巻線品質良好',     f_anno_g, p_anno_g),
]); R+=1

mw(R, 2, 18,
   '※ U軸についてもピンホールは確認されず、良好な巻線品質を確認いたしました。',
   font=f_note, align=al); R+=1
rh(R, 10); R+=1

# ──────────────────────────────────────
# ⑤ 現在の状況
# ──────────────────────────────────────
section_bar(R, '■ 現在の状況'); R+=1

mw(R, 2, 2,   'No.',       font=f_sub, fill=p_gray, align=ac, h=26)
mw(R, 3, 7,   '確認項目',  font=f_sub, fill=p_gray, align=ac, h=26)
mw(R, 8, 14,  '確認結果',  font=f_sub, fill=p_gray, align=ac, h=26)
mw(R, 15, 16, '判定',      font=f_sub, fill=p_gray, align=ac, h=26)
brd(R, 2, 16); R+=1

for no, item, result, status, bg, sf in [
    ('1', '下渡り部の品質',
     '安定的に巻線が行われており、傷・ピンホールは確認されておりません。',
     '良好', p_green_l, f_label_g),
    ('2', 'S軸(左)\n巻線部品質',
     'S軸（左）の巻線部表面にピンホールが確認されております。'
     'U軸（右）には問題はございません。現在ゲイン調整にて対応中です。',
     '調整中', p_yellow, f_label_r),
    ('3', '巻線整列',
     'プログラムの変更は行っておりませんが、目視にて良好な整列を確認しております。',
     '良好', p_green_l, f_label_g),
    ('4', 'リング\nバリ再発',
     '対策後、複数台の巻線を実施しておりますが、バリの再発は確認されておりません。'
     '引き続き経過を監視いたします。',
     '経過観察', p_yellow, f_warn),
]:
    # 最も長い列(確認結果:7列)で行高を決定
    h = calc_h(result, 7, 11, min_h=40)
    mw(R, 2, 2,   no,     font=f_body, fill=bg, align=ac, h=h)
    mw(R, 3, 7,   item,   font=f_body, fill=bg, align=ac, h=h)
    mw(R, 8, 14,  result, font=f_body, fill=bg, align=at, h=h)
    mw(R, 15, 16, status, font=sf,     fill=bg, align=ac, h=h)
    brd(R, 2, 16); R+=1

rh(R, 10); R+=1

# ──────────────────────────────────────
# ⑥ 今後の課題・懸念点
# ──────────────────────────────────────
section_bar(R, '■ 今後の課題・懸念点'); R+=1

for label, text in [
    ('懸念点1',
     'コア押上げリングのバリが再発する可能性がございます。'
     '現状では複数台巻線後もバリの再発は確認されておりませんが、定期的な点検を継続いたします。'),
    ('懸念点2',
     '今回の巻線機で改善が成功した場合、現行4台への横展開を予定しておりますが、'
     '大幅な改造が必要となる見込みです。横展開の計画については別途ご相談させていただきます。'),
]:
    h = calc_h(text, 13, 11, min_h=40)   # 内容列(col4-16 = 13列)で計算
    mw(R, 2, 3,  label, font=f_sub,  fill=p_yellow, align=ac, h=h)
    mw(R, 4, 16, text,  font=f_body, fill=p_yellow, align=at, h=h)
    brd(R, 2, 16); R+=1

rh(R, 10); R+=1

# ──────────────────────────────────────
# ⑦ 出荷スケジュール
# ──────────────────────────────────────
section_bar(R, '■ 出荷スケジュール'); R+=1

mw(R, 3, 5,   '日程',     font=f_sub, fill=p_gray, align=ac, h=26)
mw(R, 6, 10,  'イベント', font=f_sub, fill=p_gray, align=ac, h=26)
mw(R, 11, 13, '場所',     font=f_sub, fill=p_gray, align=ac, h=26)
brd(R, 3, 13); R+=1

for date, event, place, bg in [
    ('4月20日（月）', '出荷', 'テクノ',     p_sch1),
    ('5月7日',        '出港', '日本',       p_sch1),
    ('5月18日',       '入港', 'タイ',       p_sch2),
    ('5月最終週',     '到着', 'カミンブリ', p_sch2),
]:
    mw(R, 3, 5,   date,  font=f_sch,  fill=bg, align=ac, h=28)
    mw(R, 6, 10,  event, font=f_body, fill=bg, align=ac, h=28)
    mw(R, 11, 13, place, font=f_body, fill=bg, align=ac, h=28)
    brd(R, 3, 13); R+=1

rh(R, 6); R+=1
mw(R, 3, 13,
   '4/20 出荷（テクノ）→ 5/7 出港（日本）→ 5/18 入港（タイ）→ 5月末 カミンブリ到着',
   font=f_flow, align=ac, h=30); R+=1

rh(R, 10); R+=1
mw(R, 2, 16, '── 以上 ──', font=f_footer, align=ac, h=24)

# ── 印刷設定 ──
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize   = 8   # A3
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

output = 'C:/Users/SEIGI-N13/Desktop/EOP巻線機_MGワイヤ傷_報告書.xlsx'
wb.save(output)
print(f'完了（最終行: {R}）')

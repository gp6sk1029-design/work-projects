# -*- coding: utf-8 -*-
"""
【パッチ用スクリプト】θ軸の機械位置倍率を ×100 → ×10 に変更

【変更対象】
  - F列（θ機械位置→μm）：=(B{r}*100-$W$61)*$W$26  →  =(B{r}*10-$W$61)*$W$26
  - Z（H列）は ×100 のまま（変更なし）
  - ヘッダ row 3 col 2 の「(×100)」表記を「(×10)」に更新
"""

import sys
import os
import string
import re
from datetime import datetime
from openpyxl import load_workbook

DESKTOP = r"C:\Users\SEIGI-N13\Desktop"
BASE_NAME = "【巻線検証】SLM8000軌道_AISIN反映"


def find_latest_input():
    td = datetime.now().strftime('%Y%m%d')
    for letter in reversed(string.ascii_uppercase):
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if os.path.exists(p):
            return p
    raise RuntimeError("ファイルが見つかりません")


def next_output_path():
    td = datetime.now().strftime('%Y%m%d')
    for letter in string.ascii_uppercase:
        p = os.path.join(DESKTOP, f"{BASE_NAME}_{td}{letter}.xlsx")
        if not os.path.exists(p):
            return p, letter
    raise RuntimeError("使い切り")


def patch_theta_mult(wb):
    ws = wb['NAKAM']
    print(f"  NAKAM max_row={ws.max_row}")

    # F列（col 6）の *100 を *10 に置換
    # 元形式: =(B{r}*100-$W$61)*$W$26
    # 新形式: =(B{r}*10-$W$61)*$W$26
    pat = re.compile(r'\bB(\d+)\*100-\$W\$61\b')
    patched = 0
    skipped = 0

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=6)  # F列
        v = cell.value
        if v is None or not isinstance(v, str) or not v.startswith('='):
            continue
        # 既に *10 になっている場合スキップ
        if f'B{r}*10-' in v and f'B{r}*100-' not in v:
            skipped += 1
            continue

        new_v = pat.sub(lambda m: f'B{m.group(1)}*10-$W$61', v)
        if new_v != v:
            cell.value = new_v
            patched += 1

    print(f"  F列（θ機械位置→μm）を ×100 → ×10 に変更: {patched} セル")
    print(f"  スキップ（既 ×10）: {skipped} セル")

    # ─────────────────────────
    # ヘッダ更新：row 3 col 2（θ機械位置）の「(×100)」→「(×10)」
    # ─────────────────────────
    h_cell = ws.cell(row=3, column=2)
    if h_cell.value and '×100' in str(h_cell.value):
        h_cell.value = h_cell.value.replace('×100', '×10')
        print(f"  row 3 col 2 ヘッダ更新: {h_cell.value!r}")

    return patched


def main():
    input_path = find_latest_input()
    output_path, letter = next_output_path()
    print(f"入力: {input_path}")
    print(f"出力: {output_path}  （今日{letter}版）")

    wb = load_workbook(input_path)
    print(f"\n--- パッチ適用中 ---")
    patch_theta_mult(wb)

    print(f"\n--- 保存 ---")
    try:
        wb.save(output_path)
        print(f"完了: {output_path}")
    except PermissionError:
        print(f"ERROR: 出力先が開かれています: {output_path}", file=sys.stderr)
        sys.exit(1)

    # 検証
    wb2 = load_workbook(output_path)
    ws = wb2['NAKAM']
    print(f"\n--- 検証 ---")
    print(f"  row 3 col 2 ヘッダ: {ws.cell(row=3, column=2).value!r}")
    print(f"  サンプル数式（F列=θ×10, H列=Z×100）:")
    for r in (4, 100, 1000, 5000):
        f = ws.cell(row=r, column=6).value
        h = ws.cell(row=r, column=8).value
        print(f"    row{r}: F={f}  H={h}")


if __name__ == '__main__':
    main()

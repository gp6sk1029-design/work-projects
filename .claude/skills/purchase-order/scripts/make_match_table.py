# -*- coding: utf-8 -*-
"""見積書（明細JSON）と注文書Excelを1件ずつ突合し、確認用の照合表Excelを作る。

左に見積書・右に注文書を並べ、判定列を色分けする。
注文書の行順と同じ並びで出力するので、印刷して照合できる。

使い方:
    python make_match_table.py --xlsx 注文書.xlsx --sheet NTS --json items.json \
        --start-page 5 --out 照合表.xlsx [--fixed "誤った型式=正しい型式,..."]

--fixed は「見積書の誤記を注文書側で直した」ペア。判定を「修正済」として扱う。
"""
import sys, os, re, json, argparse

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLS = {"B": 2, "J": 10, "R": 18, "Z": 26, "AH": 34}
BLOCKS = [(7, 16, 40, 41), (52, 61, 85, 86), (97, 106, 130, 131)]
ROWS_PER_PAGE = 25
# 品名から取り除く加工処理表記（(裏)などの部位表記は残す）
PROC = [r"\(無電解ニッケル\)", r"\(アルマイト\)", r"\(チッカ\)", r"\(焼入れ[^)]*\)", r"※処理無し"]


def strip_proc(s):
    for p in PROC:
        s = re.sub(p, "", s or "")
    return (s or "").strip()


def read_order(path, sheet, start_page, need_pages):
    wb = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)   # ページ番号は数式のことがある
    ws, wsv = wb[sheet], wbv[sheet]
    pages = []
    for prow, r0, r1, rsum in BLOCKS:
        for cl, ci in COLS.items():
            pnum = wsv.cell(row=prow, column=ci + 1).value
            if pnum is None:
                continue
            try:
                pno = int(pnum)
            except (TypeError, ValueError):
                continue                                  # 計算値が無い/数値でない頁は飛ばす
            pages.append({"no": pno, "col": ci, "r0": r0, "label": f"{cl}{prow}"})
    pages.sort(key=lambda p: p["no"])
    if start_page:
        pages = [p for p in pages if p["no"] >= start_page]
    rows = []
    for p in pages[:need_pages]:
        for k in range(ROWS_PER_PAGE):
            r = p["r0"] + k
            rows.append({
                "page": p["no"], "row": r,
                "katashiki": ws.cell(row=r, column=p["col"]).value,
                "hinmei": ws.cell(row=r, column=p["col"] + 1).value,
                "qty": ws.cell(row=r, column=p["col"] + 3).value,
                "tanka": ws.cell(row=r, column=p["col"] + 4).value,
                "noki": ws.cell(row=r, column=p["col"] + 6).value,
            })
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--json", required=True, help="見積書の明細JSON")
    ap.add_argument("--start-page", type=int, default=None)
    ap.add_argument("--out", required=True)
    ap.add_argument("--fixed", default="", help='"555A=55A,28A=38A" 形式')
    ap.add_argument("--title", default="見積書 と 注文書 の照合表")
    a = ap.parse_args()

    items = json.load(open(a.json, encoding="utf-8"))
    need = -(-len(items) // ROWS_PER_PAGE)
    order = read_order(a.xlsx, a.sheet, a.start_page, need)
    fixed = {}
    for pair in filter(None, a.fixed.split(",")):
        k, _, v = pair.partition("=")
        fixed[k.strip()] = v.strip()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "照合表"
    thin = Side(style="thin", color="AAAAAA")
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR, HF = PatternFill("solid", fgColor="1F4E79"), Font(color="FFFFFF", bold=True, size=9)
    OK = PatternFill("solid", fgColor="C6EFCE")
    NG = PatternFill("solid", fgColor="FFC7CE")
    FIX = PatternFill("solid", fgColor="FFE699")
    G1 = PatternFill("solid", fgColor="DDEBF7")
    G2 = PatternFill("solid", fgColor="FFF2CC")

    ws["A1"] = a.title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"対象: {os.path.basename(a.xlsx)} ／ タブ「{a.sheet}」 ／ 注文書の行順と同じ並びです"
    ws["A2"].font = Font(size=9, color="666666")
    ws["A4"] = "◀ 見積書（原本の読み取り値）"
    ws["A4"].font = Font(bold=True, size=10, color="1F4E79")
    ws["H4"] = "◀ 注文書（入力結果）"
    ws["H4"].font = Font(bold=True, size=10, color="C55A11")
    head = ["No", "型式", "品名（原文）", "数量", "単価", "金額",
            "頁", "行", "型式", "品名", "数量", "単価", "金額", "納期", "判定"]
    for c, h in enumerate(head, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.fill, cell.font = HDR, HF
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BD

    ng = fx = 0
    for i, (q, o) in enumerate(zip(items, order), 1):
        kq, nq = q["katashiki"], q["hinmei"]
        qq, tq = q["qty"], q.get("tanka")
        is_fix = (kq != o["katashiki"]) and (fixed.get(kq) == o["katashiki"])
        kata_ok = (kq == o["katashiki"]) or is_fix
        ok = (kata_ok and strip_proc(nq) == (o["hinmei"] or "")
              and qq == o["qty"] and (tq or 0) == (o["tanka"] or 0))
        if is_fix:
            fx += 1
        elif not ok:
            ng += 1
        judge = "修正済" if is_fix else ("一致" if ok else "★要確認")
        vals = [i, kq, nq, qq, tq, (qq or 0) * (tq or 0),
                o["page"], o["row"], o["katashiki"], o["hinmei"], o["qty"], o["tanka"],
                (o["qty"] or 0) * (o["tanka"] or 0), o["noki"] or "", judge]
        r = 5 + i
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border, cell.font = BD, Font(size=9)
            cell.fill = G1 if c <= 6 else (G2 if c <= 14 else OK)
            if c in (4, 5, 6, 11, 12, 13):
                cell.number_format = "#,##0"
            if c == 15:
                cell.fill = FIX if is_fix else (OK if ok else NG)
                cell.font = Font(size=9, bold=True,
                                 color="7F6000" if is_fix else ("006100" if ok else "9C0006"))
                cell.alignment = Alignment(horizontal="center")
            if c == 9 and is_fix:
                cell.fill, cell.font = FIX, Font(size=9, bold=True, color="7F6000")
            if c == 14 and v:
                cell.font = Font(size=9, bold=True, color="C55A11")

    last = 5 + len(items)
    ws.cell(row=last + 1, column=1, value="合計").font = Font(bold=True)
    for col, letter in ((6, "F"), (13, "M")):
        c = ws.cell(row=last + 1, column=col, value=f"=SUM({letter}6:{letter}{last})")
        c.number_format, c.font, c.fill = "#,##0", Font(bold=True), OK
    ws.cell(row=last + 3, column=1,
            value=f"◆ 判定：{len(items)}件中 一致{len(items)-ng-fx}件 / 修正済{fx}件 / 要確認{ng}件"
            ).font = Font(bold=True, size=11, color="006100" if ng == 0 else "9C0006")
    if fixed:
        ws.cell(row=last + 4, column=1,
                value="◆ 型式を修正した箇所: " + ", ".join(f"{k}→{v}" for k, v in fixed.items())
                ).font = Font(size=10, color="7F6000")

    for col, w in zip("ABCDEFGHIJKLMNO", (5, 16, 30, 6, 9, 10, 5, 5, 16, 26, 6, 9, 10, 11, 9)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:O{last}"
    wb.save(a.out)
    print(f"一致{len(items)-ng-fx}件 / 修正済{fx}件 / 要確認{ng}件")
    print("保存:", a.out)
    return 0 if ng == 0 else 1


if __name__ == "__main__":
    sys.exit(main())

# -*- coding: utf-8 -*-
"""注文書Excelの構造（ページ番号・注文No.・使用状況）を一覧表示する。

使い方:
    python inspect_order_sheet.py <注文書.xlsx> [タブ名]
タブ名を省略すると全タブのサマリーを出す。
"""
import sys, os

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

# 1シート内のページ配置（型式列インデックス）と行ブロック
COLS = {"B": 2, "J": 10, "R": 18, "Z": 26, "AH": 34}
BLOCKS = [(7, 16, 40, 41), (52, 61, 85, 86), (97, 106, 130, 131)]  # (頁番号行, 明細開始, 明細終了, 合計行)


def inspect(path, sheet):
    wb = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    ws, wsv = wb[sheet], wbv[sheet]
    print(f"\n=== タブ「{sheet}」 dimensions={ws.dimensions} ===")
    print("宛先:", ws["B4"].value)
    print(f"{'位置':<8}{'頁':<4}{'注文No.':<18}{'使用':<6}{'合計':>12}  先頭データ")
    empty_pages = []
    for prow, r0, r1, rsum in BLOCKS:
        for cl, ci in COLS.items():
            pnum = wsv.cell(row=prow, column=ci + 1).value
            ono = wsv.cell(row=prow + 1, column=ci + 1).value
            if pnum is None and ono is None:
                continue
            used, first = 0, ""
            for r in range(r0, r1 + 1):
                v = ws.cell(row=r, column=ci).value
                if v not in (None, ""):
                    used += 1
                    if not first:
                        first = str(v)[:24]
            total = wsv.cell(row=rsum, column=ci + 5).value or 0
            print(f"{cl}{prow:<7}{str(pnum):<4}{str(ono):<18}{used:<6}{total:>12,.0f}  {first}")
            if used == 0:
                empty_pages.append((str(ono), cl, prow, r0, r1))
    if empty_pages:
        print("\n空きページ（ここから入力できます）:")
        for ono, cl, prow, r0, r1 in empty_pages[:5]:
            print(f"  {ono}  型式列={cl}  明細行={r0}〜{r1}（{r1-r0+1}行）")
    return empty_pages


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    path = os.path.abspath(sys.argv[1])
    lock = os.path.join(os.path.dirname(path), "~$" + os.path.basename(path))
    if os.path.exists(lock):
        print("⚠️ このファイルはExcelで開かれています。閉じてから編集してください。")
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    if len(sys.argv) >= 3:
        inspect(path, sys.argv[2])
    else:
        print("タブ一覧:", sheets)
        for s in sheets:
            try:
                inspect(path, s)
            except Exception as e:
                print(f"  {s}: 読めません（{e}）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
